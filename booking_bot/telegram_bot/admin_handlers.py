import csv
import logging
import re
from datetime import date, timedelta
from math import ceil
from typing import Dict, List, Optional, Tuple
from io import BytesIO, StringIO

from django.db.models import (
    Sum,
    F,
    Avg,
    Count,
    ExpressionWrapper,
    DurationField,
)
from telegram import KeyboardButton, ReplyKeyboardMarkup
from openpyxl import Workbook

from booking_bot.users.models import UserProfile, RealEstateAgency
from booking_bot.listings.models import Property, City, District, PropertyPhoto
from booking_bot.bookings.models import Booking
from .constants import (
    STATE_ADMIN_ADD_PROPERTY,
    STATE_ADMIN_ADD_DESC,
    STATE_ADMIN_ADD_ADDRESS,
    STATE_ADMIN_ADD_CITY,
    STATE_ADMIN_ADD_DISTRICT,
    STATE_ADMIN_ADD_CLASS,
    STATE_ADMIN_ADD_ROOMS,
    STATE_ADMIN_ADD_AREA,
    STATE_ADMIN_ADD_PRICE,
    STATE_ADMIN_ADD_PHOTOS,
    STATE_SEARCH_REFINED,
    _get_profile,
    log_handler,
    log_state_transition,
    start_command_handler, User,
    STATE_EDIT_PROPERTY_MENU,
    STATE_WAITING_NEW_PRICE,
    STATE_WAITING_NEW_DESCRIPTION,
    STATE_WAITING_NEW_STATUS, PAGE_SIZE, STATE_PHOTO_MANAGEMENT,
)

from .utils import send_telegram_message
from ..settings import TELEGRAM_BOT_TOKEN

logger = logging.getLogger(__name__)

ANALYTICS_PAGE_SIZE = 5
PERIOD_PRESETS: Dict[str, Dict[str, object]] = {
    "day": {"days": 0, "label": "день"},
    "week": {"days": 7, "label": "неделю"},
    "month": {"days": 30, "label": "месяц"},
    "quarter": {"days": 90, "label": "квартал"},
    "year": {"days": 365, "label": "год"},
}
PERIOD_BUTTONS = {
    "День": "day",
    "Неделя": "week",
    "Месяц": "month",
    "Квартал": "quarter",
    "Год": "year",
}
ANALYTIC_STATUSES = ["confirmed", "completed"]


def _resolve_period_bounds(period: str) -> Tuple[date, date, str]:
    preset = PERIOD_PRESETS.get(period, PERIOD_PRESETS["month"])
    today = date.today()
    days = int(preset["days"])
    start = today - timedelta(days=days) if days else today
    return start, today, str(preset["label"])


def _period_keyboard() -> List[List[KeyboardButton]]:
    return [
        [KeyboardButton("День"), KeyboardButton("Неделя"), KeyboardButton("Месяц")],
        [KeyboardButton("Квартал"), KeyboardButton("Год")],
    ]


def _slice_page(items: List[dict], page: int, page_size: int = ANALYTICS_PAGE_SIZE) -> List[dict]:
    start_index = max(page - 1, 0) * page_size
    end_index = start_index + page_size
    return items[start_index:end_index]


def _collect_realtor_metrics(period: str):
    start, end, label = _resolve_period_bounds(period)
    admins = (
        UserProfile.objects.filter(role=UserProfile.ROLE_ADMIN)
        .select_related("user", "agency")
        .order_by("user__username")
    )

    admin_ids = [profile.user_id for profile in admins if profile.user_id]
    if not admin_ids:
        return [], start, end, label

    bookings = Booking.objects.filter(
        status__in=ANALYTIC_STATUSES,
        start_date__gte=start,
        start_date__lte=end,
        property__owner_id__in=admin_ids,
    )

    owner_metrics: Dict[int, Dict[str, float]] = {
        row["property__owner_id"]: {
            "revenue": row.get("total_revenue") or 0,
            "bookings": row.get("bookings_count") or 0,
        }
        for row in bookings.values("property__owner_id").annotate(
            total_revenue=Sum("total_price"),
            bookings_count=Count("id"),
        )
    }

    property_counts: Dict[int, int] = {
        row["owner_id"]: row["property_count"]
        for row in Property.objects.filter(owner_id__in=admin_ids)
        .values("owner_id")
        .annotate(property_count=Count("id"))
    }

    results: List[Dict[str, object]] = []
    for profile in admins:
        user = profile.user
        if not user:
            continue
        display_name = user.get_full_name() or getattr(user, "username", "") or f"ID {user.id}"
        metrics = owner_metrics.get(profile.user_id, {"revenue": 0, "bookings": 0})
        results.append(
            {
                "profile": profile,
                "name": display_name,
                "agency": profile.agency.name if profile.agency else None,
                "revenue": metrics["revenue"],
                "bookings": metrics["bookings"],
                "properties": property_counts.get(profile.user_id, 0),
            }
        )

    results.sort(key=lambda item: (-float(item["revenue"]), item["name"].lower()))
    return results, start, end, label


def _collect_agency_metrics(period: str):
    start, end, label = _resolve_period_bounds(period)
    agency_members = (
        UserProfile.objects.filter(role=UserProfile.ROLE_ADMIN, agency__isnull=False)
        .select_related("agency", "user")
    )

    if not agency_members.exists():
        return [], start, end, label

    agency_member_counts = {
        row["agency_id"]: row["member_count"]
        for row in agency_members.values("agency_id").annotate(member_count=Count("id"))
    }

    agency_ids = list(agency_member_counts.keys())
    agencies = RealEstateAgency.objects.filter(id__in=agency_ids).order_by("name")

    bookings = Booking.objects.filter(
        status__in=ANALYTIC_STATUSES,
        start_date__gte=start,
        start_date__lte=end,
        property__owner__profile__agency_id__in=agency_ids,
    )

    agency_metrics: Dict[int, Dict[str, float]] = {
        row["property__owner__profile__agency_id"]: {
            "revenue": row.get("total_revenue") or 0,
            "bookings": row.get("bookings_count") or 0,
        }
        for row in bookings.values("property__owner__profile__agency_id").annotate(
            total_revenue=Sum("total_price"),
            bookings_count=Count("id"),
        )
    }

    property_counts: Dict[int, int] = {
        row["owner__profile__agency_id"]: row["property_count"]
        for row in Property.objects.filter(owner__profile__agency_id__in=agency_ids)
        .values("owner__profile__agency_id")
        .annotate(property_count=Count("id"))
    }

    results: List[Dict[str, object]] = []
    for agency in agencies:
        metrics = agency_metrics.get(agency.id, {"revenue": 0, "bookings": 0})
        results.append(
            {
                "agency": agency,
                "revenue": metrics["revenue"],
                "bookings": metrics["bookings"],
                "properties": property_counts.get(agency.id, 0),
                "members": agency_member_counts.get(agency.id, 0),
            }
        )

    results.sort(key=lambda item: (-float(item["revenue"]), item["agency"].name.lower()))
    return results, start, end, label


def _collect_agency_detail_metrics(agency: RealEstateAgency, period: str):
    start, end, label = _resolve_period_bounds(period)
    properties = Property.objects.filter(owner__profile__agency=agency)
    property_ids = list(properties.values_list("id", flat=True))

    bookings = Booking.objects.filter(
        status__in=ANALYTIC_STATUSES,
        start_date__gte=start,
        start_date__lte=end,
        property_id__in=property_ids,
    )

    summary = bookings.aggregate(
        total_revenue=Sum("total_price"),
        total_bookings=Count("id"),
    )

    cancelled = Booking.objects.filter(
        status="cancelled",
        start_date__gte=start,
        start_date__lte=end,
        property_id__in=property_ids,
    )

    duration_expr = ExpressionWrapper(
        F("end_date") - F("start_date"), output_field=DurationField()
    )
    occupied_delta = bookings.annotate(duration=duration_expr).aggregate(total=Sum("duration"))["total"]
    occupied_days = occupied_delta.days if occupied_delta else 0
    period_days = max((end - start).days or 1, 1)
    inventory_days = max(len(property_ids), 1) * period_days
    occupancy_rate = (occupied_days / inventory_days * 100) if inventory_days else 0

    cancel_breakdown = [
        (
            row["cancel_reason"] or "",
            row["total"],
        )
        for row in cancelled.values("cancel_reason").annotate(total=Count("id"))
        if row["total"]
    ]

    top_props_revenue = list(
        bookings.values("property__name")
        .annotate(total=Sum("total_price"))
        .order_by("-total")[:5]
    )
    top_props_bookings = list(
        bookings.values("property__name")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )
    top_users_count = list(
        bookings.values("user__username")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )
    top_users_spend = list(
        bookings.values("user__username")
        .annotate(total=Sum("total_price"))
        .order_by("-total")[:5]
    )
    top_realtor_revenue = list(
        bookings.values("property__owner__username")
        .annotate(total=Sum("total_price"))
        .order_by("-total")[:5]
    )
    top_realtor_bookings = list(
        bookings.values("property__owner__username")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )

    return {
        "start": start,
        "end": end,
        "label": label,
        "summary": summary,
        "properties": len(property_ids),
        "members": agency.members.count(),
        "occupancy": occupancy_rate,
        "cancel_breakdown": cancel_breakdown,
        "top_props_revenue": top_props_revenue,
        "top_props_bookings": top_props_bookings,
        "top_users_count": top_users_count,
        "top_users_spend": top_users_spend,
        "top_realtor_revenue": top_realtor_revenue,
        "top_realtor_bookings": top_realtor_bookings,
        "cancelled_total": cancelled.count(),
    }


@log_handler
def show_realtor_statistics(chat_id: int, period: str = "month", page: int = 1):
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    data, start, end, label = _collect_realtor_metrics(period)

    profile_state = {
        "state": "super_admin_realtor_stats",
        "period": period,
        "page": page,
        "analytics_scope": "realtors",
    }

    if not data:
        keyboard = _period_keyboard()
        keyboard.append([KeyboardButton("🏢 Агентства")])
        keyboard.append([KeyboardButton("🧭 Главное меню")])
        send_telegram_message(
            chat_id,
            "📊 *Риелторы*\n\nДанных за выбранный период пока нет.",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
        )
        profile.telegram_state = profile_state
        profile.save()
        return

    total = len(data)
    total_pages = max(1, ceil(total / ANALYTICS_PAGE_SIZE))
    page = max(1, min(page, total_pages))
    page_items = _slice_page(data, page)

    start_idx = (page - 1) * ANALYTICS_PAGE_SIZE
    lines = [
        f"📊 *Риелторы — доход за {label}*",
        f"Период: {start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}",
        f"Страница {page}/{total_pages}\n",
    ]

    for offset, item in enumerate(page_items, start=1):
        idx = start_idx + offset
        revenue = float(item["revenue"] or 0)
        bookings_count = int(item["bookings"] or 0)
        property_count = int(item["properties"] or 0)
        line = (
            f"{idx}. {item['name']} — {revenue:,.0f} ₸ | броней: {bookings_count} | объектов: {property_count}"
        )
        if item["agency"]:
            line += f" | агентство: {item['agency']}"
        lines.append(line)

    lines.append("\nИспользуйте кнопки для смены периода или экспорта отчета.")

    nav_row: List[KeyboardButton] = []
    if page > 1:
        nav_row.append(KeyboardButton(f"⬅️ Назад (стр. {page - 1})"))
    nav_row.append(KeyboardButton(f"📄 {page}/{total_pages}"))
    if page < total_pages:
        nav_row.append(KeyboardButton(f"➡️ Далее (стр. {page + 1})"))

    keyboard: List[List[KeyboardButton]] = []
    if nav_row:
        keyboard.append(nav_row)
    keyboard.extend(_period_keyboard())
    keyboard.append([KeyboardButton("🏢 Агентства"), KeyboardButton("📈 Экспорт XLSX")])
    keyboard.append([KeyboardButton("🧭 Главное меню")])

    send_telegram_message(
        chat_id,
        "\n".join(lines),
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )

    profile_state["page"] = page
    profile.telegram_state = profile_state
    profile.save()


@log_handler
def show_agency_statistics(chat_id: int, period: str = "month", page: int = 1):
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    data, start, end, label = _collect_agency_metrics(period)

    state_data = {
        "state": "super_admin_agency_list",
        "period": period,
        "page": page,
        "analytics_scope": "agency_list",
        "agency_lookup": {},
    }

    if not data:
        keyboard = _period_keyboard()
        keyboard.append([KeyboardButton("📊 Риелторы")])
        keyboard.append([KeyboardButton("🧭 Главное меню")])
        send_telegram_message(
            chat_id,
            "🏢 *Агентства*\n\nДанных за выбранный период пока нет.",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
        )
        profile.telegram_state = state_data
        profile.save()
        return

    total = len(data)
    total_pages = max(1, ceil(total / ANALYTICS_PAGE_SIZE))
    page = max(1, min(page, total_pages))
    page_items = _slice_page(data, page)

    lines = [
        f"🏢 *Агентства — доход за {label}*",
        f"Период: {start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}",
        f"Страница {page}/{total_pages}\n",
    ]

    start_idx = (page - 1) * ANALYTICS_PAGE_SIZE
    keyboard: List[List[KeyboardButton]] = []
    for offset, item in enumerate(page_items, start=1):
        idx = start_idx + offset
        agency = item["agency"]
        revenue = float(item["revenue"] or 0)
        bookings_count = int(item["bookings"] or 0)
        property_count = int(item["properties"] or 0)
        member_count = int(item["members"] or 0)
        lines.append(
            f"{idx}. {agency.name} — {revenue:,.0f} ₸ | броней: {bookings_count} | объектов: {property_count} | риелторов: {member_count}"
        )
        button_text = f"🏢 {agency.name} • {revenue:,.0f} ₸"
        state_data["agency_lookup"][button_text] = agency.id
        keyboard.append([KeyboardButton(button_text)])

    lines.append("\nВыберите агентство для подробной аналитики или измените период.")

    nav_row: List[KeyboardButton] = []
    if page > 1:
        nav_row.append(KeyboardButton(f"⬅️ Назад (стр. {page - 1})"))
    nav_row.append(KeyboardButton(f"📄 {page}/{total_pages}"))
    if page < total_pages:
        nav_row.append(KeyboardButton(f"➡️ Далее (стр. {page + 1})"))

    if nav_row:
        keyboard.append(nav_row)
    keyboard.extend(_period_keyboard())
    keyboard.append([KeyboardButton("📊 Риелторы"), KeyboardButton("📈 Экспорт XLSX")])
    keyboard.append([KeyboardButton("🧭 Главное меню")])

    send_telegram_message(
        chat_id,
        "\n".join(lines),
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )

    state_data["page"] = page
    profile.telegram_state = state_data
    profile.save()


@log_handler
def show_agency_details(
    chat_id: int,
    agency_id: int,
    period: str = "month",
    source_page: Optional[int] = None,
):
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    summary_period: Optional[Tuple[date, date]] = None

    try:
        agency = RealEstateAgency.objects.get(id=agency_id)
    except RealEstateAgency.DoesNotExist:
        send_telegram_message(chat_id, "Агентство не найдено.")
        return

    metrics = _collect_agency_detail_metrics(agency, period)
    summary = metrics.get("summary") or {"total_revenue": 0, "total_bookings": 0}
    start = metrics["start"]
    end = metrics["end"]
    label = metrics["label"]

    lines = [
        f"🏢 *{agency.name}* — аналитика за {label}",
        f"Период: {start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}",
        (
            f"Доход: {float(summary.get('total_revenue') or 0):,.0f} ₸ | "
            f"Брони: {int(summary.get('total_bookings') or 0)}"
        ),
        (
            f"Объектов: {metrics['properties']} | Риелторов: {metrics['members']} | "
            f"Загрузка: {metrics['occupancy']:.1f}% | Отмен: {metrics['cancelled_total']}"
        ),
        "",
    ]

    def _format_ranked(items: List[dict], value_key: str, label_key: str, is_currency: bool = False):
        if not items:
            return ["— Нет данных —"]
        formatted = []
        for idx, row in enumerate(items, start=1):
            label_value = row.get(label_key) or "—"
            metric_value = row.get(value_key) or 0
            if is_currency:
                metric_value = f"{float(metric_value):,.0f} ₸"
            else:
                metric_value = str(metric_value)
            formatted.append(f"{idx}. {label_value}: {metric_value}")
        return formatted

    lines.append("🏠 Топ-5 квартир по доходу:")
    lines.extend(_format_ranked(metrics["top_props_revenue"], "total", "property__name", is_currency=True))
    lines.append("")
    lines.append("🏠 Топ-5 квартир по бронированиям:")
    lines.extend(_format_ranked(metrics["top_props_bookings"], "count", "property__name"))
    lines.append("")
    lines.append("👥 Топ-5 гостей по количеству заселений:")
    lines.extend(_format_ranked(metrics["top_users_count"], "count", "user__username"))
    lines.append("")
    lines.append("💸 Топ-5 гостей по сумме трат:")
    lines.extend(_format_ranked(metrics["top_users_spend"], "total", "user__username", is_currency=True))
    lines.append("")
    lines.append("🧑‍💼 Топ-5 риелторов по доходу:")
    lines.extend(_format_ranked(metrics["top_realtor_revenue"], "total", "property__owner__username", is_currency=True))
    lines.append("")
    lines.append("🧑‍💼 Топ-5 риелторов по бронированиям:")
    lines.extend(_format_ranked(metrics["top_realtor_bookings"], "count", "property__owner__username"))
    lines.append("")

    cancel_labels = dict(Booking.CANCEL_REASON_CHOICES)
    lines.append("🚫 Отмены по причинам:")
    if not metrics["cancel_breakdown"]:
        lines.append("— Нет данных —")
    else:
        for reason_code, total in metrics["cancel_breakdown"]:
            reason_label = cancel_labels.get(reason_code, reason_code or "Без причины")
            lines.append(f"• {reason_label}: {total}")

    keyboard = _period_keyboard()
    keyboard.insert(0, [KeyboardButton("⬅️ К списку агентств")])
    keyboard.append([KeyboardButton("📊 Риелторы"), KeyboardButton("🏢 Агентства")])
    keyboard.append([KeyboardButton("📈 Экспорт XLSX")])
    keyboard.append([KeyboardButton("🧭 Главное меню")])

    send_telegram_message(
        chat_id,
        "\n".join(lines),
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )

    profile.telegram_state = {
        "state": "super_admin_agency_detail",
        "period": period,
        "analytics_scope": "agency_detail",
        "agency_id": agency.id,
        "previous_page": source_page,
    }
    profile.save()


# Новые состояния для кодов доступа
STATE_ADMIN_ADD_ENTRY_FLOOR = "admin_add_entry_floor"
STATE_ADMIN_ADD_ENTRY_CODE = "admin_add_entry_code"
STATE_ADMIN_ADD_KEY_SAFE = "admin_add_key_safe"
STATE_ADMIN_ADD_OWNER_PHONE = "admin_add_owner_phone"
STATE_ADMIN_ADD_INSTRUCTIONS = "admin_add_instructions"


@log_handler
def handle_add_property_start(chat_id: int, text: str) -> Optional[bool]:
    profile = _get_profile(chat_id)
    state_data = profile.telegram_state or {}
    state = state_data.get("state")

    admin_states = {
        STATE_ADMIN_ADD_PROPERTY,
        STATE_ADMIN_ADD_DESC,
        STATE_ADMIN_ADD_ADDRESS,
        STATE_ADMIN_ADD_CITY,
        STATE_ADMIN_ADD_DISTRICT,
        STATE_ADMIN_ADD_CLASS,
        STATE_ADMIN_ADD_ROOMS,
        STATE_ADMIN_ADD_AREA,
        STATE_ADMIN_ADD_PRICE,
        STATE_ADMIN_ADD_ENTRY_FLOOR,
        STATE_ADMIN_ADD_ENTRY_CODE,
        STATE_ADMIN_ADD_KEY_SAFE,
        STATE_ADMIN_ADD_OWNER_PHONE,
        STATE_ADMIN_ADD_INSTRUCTIONS,
        STATE_ADMIN_ADD_PHOTOS,
    }

    # Триггер на первый шаг
    if text == "➕ Добавить квартиру" and state not in admin_states:
        if profile.role not in ("admin", "super_admin", "super_user"):
            send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
            return True
        jwt = (state_data or {}).get("jwt_access_token")
        new_state = {"state": STATE_ADMIN_ADD_PROPERTY, "new_property": {}}
        if jwt:
            new_state["jwt_access_token"] = jwt
        profile.telegram_state = new_state
        profile.save()
        rm = ReplyKeyboardMarkup(
            [[KeyboardButton("❌ Отмена")]],
            resize_keyboard=True,
            input_field_placeholder="Например: Уютная студия",
        ).to_dict()
        send_telegram_message(
            chat_id,
            "➕ *Добавление новой квартиры*\n\n"
            "Шаг 1/15: Введите *название* квартиры:",
            reply_markup=rm,
        )
        return True

    if state not in admin_states:
        return False

    # Отмена в любой момент
    if text == "❌ Отмена":
        profile.telegram_state = {}
        profile.save()
        start_command_handler(chat_id)
        return True

    # 1→2: Название → Описание
    if state == STATE_ADMIN_ADD_PROPERTY:
        state_data["new_property"]["name"] = text.strip()
        state_data["state"] = STATE_ADMIN_ADD_DESC
        profile.telegram_state = state_data
        profile.save()
        rm = ReplyKeyboardMarkup(
            [[KeyboardButton("❌ Отмена")]],
            resize_keyboard=True,
            input_field_placeholder="Введите описание",
        ).to_dict()
        send_telegram_message(
            chat_id, "Шаг 2/15: Введите *описание* квартиры:", reply_markup=rm
        )
        return True

    # 2→3: Описание → Адрес
    if state == STATE_ADMIN_ADD_DESC:
        state_data["new_property"]["description"] = text.strip()
        state_data["state"] = STATE_ADMIN_ADD_ADDRESS
        profile.telegram_state = state_data
        profile.save()
        rm = ReplyKeyboardMarkup(
            [[KeyboardButton("❌ Отмена")]],
            resize_keyboard=True,
            input_field_placeholder="Введите адрес",
        ).to_dict()
        send_telegram_message(
            chat_id, "Шаг 3/15: Введите *адрес* квартиры:", reply_markup=rm
        )
        return True

    # 3→4: Адрес → Город
    if state == STATE_ADMIN_ADD_ADDRESS:
        state_data["new_property"]["address"] = text.strip()
        state_data["state"] = STATE_ADMIN_ADD_CITY
        profile.telegram_state = state_data
        profile.save()
        cities = City.objects.all().order_by("name")
        kb = [[KeyboardButton(c.name)] for c in cities]
        kb.append([KeyboardButton("❌ Отмена")])
        rm = ReplyKeyboardMarkup(
            kb, resize_keyboard=True, input_field_placeholder="Выберите город"
        ).to_dict()
        send_telegram_message(chat_id, "Шаг 4/15: Выберите *город*:", reply_markup=rm)
        return True

    # 4→5: Город → Район
    if state == STATE_ADMIN_ADD_CITY:
        try:
            city = City.objects.get(name=text)
            state_data["new_property"]["city_id"] = city.id
            state_data["state"] = STATE_ADMIN_ADD_DISTRICT
            profile.telegram_state = state_data
            profile.save()
            districts = District.objects.filter(city=city).order_by("name")
            kb = [[KeyboardButton(d.name)] for d in districts]
            kb.append([KeyboardButton("❌ Отмена")])
            rm = ReplyKeyboardMarkup(
                kb, resize_keyboard=True, input_field_placeholder="Выберите район"
            ).to_dict()
            send_telegram_message(
                chat_id, f"Шаг 5/15: Выберите *район* в {city.name}:", reply_markup=rm
            )
        except City.DoesNotExist:
            send_telegram_message(chat_id, "Город не найден. Попробуйте ещё раз.")
        return True

    # 5→6: Район → Класс
    if state == STATE_ADMIN_ADD_DISTRICT:
        try:
            district = District.objects.get(
                name=text, city_id=state_data["new_property"]["city_id"]
            )
            state_data["new_property"]["district_id"] = district.id
            state_data["state"] = STATE_ADMIN_ADD_CLASS
            profile.telegram_state = state_data
            profile.save()
            classes = [
                ("comfort", "Комфорт"),
                ("business", "Бизнес"),
                ("premium", "Премиум"),
            ]
            kb = [[KeyboardButton(lbl)] for _, lbl in classes]
            kb.append([KeyboardButton("❌ Отмена")])
            rm = ReplyKeyboardMarkup(
                kb, resize_keyboard=True, input_field_placeholder="Выберите класс"
            ).to_dict()
            send_telegram_message(
                chat_id, "Шаг 6/15: Выберите *класс* жилья:", reply_markup=rm
            )
        except District.DoesNotExist:
            send_telegram_message(chat_id, "Район не найден. Попробуйте ещё раз.")
        return True

    # 6→7: Класс → Комнаты
    if state == STATE_ADMIN_ADD_CLASS:
        mapping = {"Комфорт": "comfort", "Бизнес": "business", "Премиум": "premium"}
        if text in mapping:
            state_data["new_property"]["property_class"] = mapping[text]
            state_data["state"] = STATE_ADMIN_ADD_ROOMS
            profile.telegram_state = state_data
            profile.save()
            kb = [[KeyboardButton(str(n))] for n in [1, 2, 3, "4+"]]
            kb.append([KeyboardButton("❌ Отмена")])
            rm = ReplyKeyboardMarkup(
                kb, resize_keyboard=True, input_field_placeholder="Сколько комнат?"
            ).to_dict()
            send_telegram_message(
                chat_id, "Шаг 7/15: Сколько *комнат*?", reply_markup=rm
            )
        else:
            send_telegram_message(chat_id, "Неверный выбор. Попробуйте ещё раз.")
        return True

    # 7→8: Комнаты → Площадь
    if state == STATE_ADMIN_ADD_ROOMS:
        try:
            rooms = 4 if text == "4+" else int(text)
            state_data["new_property"]["number_of_rooms"] = rooms
            state_data["state"] = STATE_ADMIN_ADD_AREA
            profile.telegram_state = state_data
            profile.save()
            rm = ReplyKeyboardMarkup(
                [[KeyboardButton("❌ Отмена")]],
                resize_keyboard=True,
                input_field_placeholder="Введите площадь",
            ).to_dict()
            send_telegram_message(
                chat_id, "Шаг 8/15: Введите *площадь* (м²):", reply_markup=rm
            )
        except ValueError:
            send_telegram_message(
                chat_id, "Неверный формат. Выберите количество комнат."
            )
        return True

    # 8→9: Площадь → Цена
    if state == STATE_ADMIN_ADD_AREA:
        try:
            area = float(text.replace(",", "."))
            state_data["new_property"]["area"] = area
            state_data["state"] = STATE_ADMIN_ADD_PRICE
            profile.telegram_state = state_data
            profile.save()
            rm = ReplyKeyboardMarkup(
                [[KeyboardButton("❌ Отмена")]],
                resize_keyboard=True,
                input_field_placeholder="Введите цену",
            ).to_dict()
            send_telegram_message(
                chat_id, "Шаг 9/15: Введите *цену* за сутки (₸):", reply_markup=rm
            )
        except ValueError:
            send_telegram_message(chat_id, "Неверный формат площади. Введите число.")
        return True

    # 9→10: Цена → Этаж
    if state == STATE_ADMIN_ADD_PRICE:
        try:
            price = float(text.replace(",", "."))
            state_data["new_property"]["price_per_day"] = price
            state_data["state"] = STATE_ADMIN_ADD_ENTRY_FLOOR
            profile.telegram_state = state_data
            profile.save()
            rm = ReplyKeyboardMarkup(
                [[KeyboardButton("Пропустить")], [KeyboardButton("❌ Отмена")]],
                resize_keyboard=True,
                input_field_placeholder="Введите этаж",
            ).to_dict()
            send_telegram_message(
                chat_id,
                "Шаг 10/15: Введите *этаж* квартиры или нажмите 'Пропустить':",
                reply_markup=rm,
            )
        except ValueError:
            send_telegram_message(chat_id, "Неверный формат цены. Введите число.")
        return True

    # 10→11: Этаж → Код домофона
    if state == STATE_ADMIN_ADD_ENTRY_FLOOR:
        if text != "Пропустить":
            try:
                floor = int(text)
                state_data["new_property"]["entry_floor"] = floor
            except ValueError:
                send_telegram_message(
                    chat_id,
                    "Неверный формат этажа. Введите число или нажмите 'Пропустить'.",
                )
                return True

        state_data["state"] = STATE_ADMIN_ADD_ENTRY_CODE
        profile.telegram_state = state_data
        profile.save()
        rm = ReplyKeyboardMarkup(
            [[KeyboardButton("Пропустить")], [KeyboardButton("❌ Отмена")]],
            resize_keyboard=True,
            input_field_placeholder="Введите код домофона",
        ).to_dict()
        send_telegram_message(
            chat_id,
            "Шаг 11/15: Введите *код домофона* или нажмите 'Пропустить':",
            reply_markup=rm,
        )
        return True

    # 11→12: Код домофона → Код сейфа
    if state == STATE_ADMIN_ADD_ENTRY_CODE:
        if text != "Пропустить":
            state_data["new_property"]["entry_code"] = text.strip()

        state_data["state"] = STATE_ADMIN_ADD_KEY_SAFE
        profile.telegram_state = state_data
        profile.save()
        rm = ReplyKeyboardMarkup(
            [[KeyboardButton("Пропустить")], [KeyboardButton("❌ Отмена")]],
            resize_keyboard=True,
            input_field_placeholder="Введите код сейфа с ключами",
        ).to_dict()
        send_telegram_message(
            chat_id,
            "Шаг 12/15: Введите *код сейфа с ключами* или нажмите 'Пропустить':",
            reply_markup=rm,
        )
        return True

    # 12→13: Код сейфа → Телефон владельца
    if state == STATE_ADMIN_ADD_KEY_SAFE:
        if text != "Пропустить":
            state_data["new_property"]["key_safe_code"] = text.strip()

        state_data["state"] = STATE_ADMIN_ADD_OWNER_PHONE
        profile.telegram_state = state_data
        profile.save()
        rm = ReplyKeyboardMarkup(
            [[KeyboardButton("Пропустить")], [KeyboardButton("❌ Отмена")]],
            resize_keyboard=True,
            input_field_placeholder="+7 XXX XXX XX XX",
        ).to_dict()
        send_telegram_message(
            chat_id,
            "Шаг 12/15: Введите *код сейфа с ключами или код от двери* или нажмите 'Пропустить':",
            reply_markup=rm,
        )
        return True

    # 13→14: Телефон → Инструкции
    if state == STATE_ADMIN_ADD_OWNER_PHONE:
        if text != "Пропустить":
            state_data["new_property"]["owner_phone"] = text.strip()

        state_data["state"] = STATE_ADMIN_ADD_INSTRUCTIONS
        profile.telegram_state = state_data
        profile.save()
        rm = ReplyKeyboardMarkup(
            [[KeyboardButton("Пропустить")], [KeyboardButton("❌ Отмена")]],
            resize_keyboard=True,
            input_field_placeholder="Введите инструкции по заселению",
        ).to_dict()
        send_telegram_message(
            chat_id,
            "Шаг 14/15: Введите *инструкции по заселению* (как найти квартиру, особенности) или нажмите 'Пропустить':",
            reply_markup=rm,
        )
        return True

    # 14→15: Инструкции → Создание и фото
    if state == STATE_ADMIN_ADD_INSTRUCTIONS:
        if text != "Пропустить":
            state_data["new_property"]["entry_instructions"] = text.strip()

        # Создаем квартиру в БД
        try:
            np = state_data["new_property"]
            prop = Property.objects.create(
                name=np["name"],
                description=np["description"],
                address=np["address"],
                district_id=np["district_id"],
                property_class=np["property_class"],
                number_of_rooms=np["number_of_rooms"],
                area=np["area"],
                price_per_day=np["price_per_day"],
                entry_floor=np.get("entry_floor"),
                entry_code=np.get("entry_code"),
                key_safe_code=np.get("key_safe_code"),
                owner_phone=np.get("owner_phone"),
                entry_instructions=np.get("entry_instructions"),
                owner=profile.user,
                status="Свободна",
            )

            state_data["new_property"]["id"] = prop.id
            state_data["state"] = STATE_ADMIN_ADD_PHOTOS
            state_data.pop("photo_mode", None)
            profile.telegram_state = state_data
            profile.save()

            # Предлагаем выбор способа загрузки фото
            rm = ReplyKeyboardMarkup(
                [
                    [KeyboardButton("📎 Отправить по URL")],
                    [KeyboardButton("📷 Загрузить фото с устройства")],
                    [KeyboardButton("⏭️ Пропустить фото")],
                    [KeyboardButton("❌ Отмена")],
                ],
                resize_keyboard=True,
                input_field_placeholder="Выберите способ",
            ).to_dict()
            send_telegram_message(
                chat_id,
                "Шаг 15/15: Выберите способ добавления фотографий:",
                reply_markup=rm,
            )
        except Exception as e:
            logger.error(f"Error creating property: {e}", exc_info=True)
            send_telegram_message(chat_id, "Ошибка при сохранении. Попробуйте снова.")
        return True

    # 15: добавление фотографий
    if state == STATE_ADMIN_ADD_PHOTOS:
        prop_id = state_data["new_property"].get("id")
        if not prop_id:
            send_telegram_message(chat_id, "Не удалось найти созданную квартиру.")
            profile.telegram_state = {}
            profile.save()
            return True

        photo_mode = state_data.get("photo_mode")

        # Пользователь выбирает способ загрузки
        if photo_mode is None:
            if text == "📎 Отправить по URL":
                state_data["photo_mode"] = "url"
                profile.telegram_state = state_data
                profile.save()
                rm = ReplyKeyboardMarkup(
                    [[KeyboardButton("✅ Завершить")], [KeyboardButton("❌ Отмена")]],
                    resize_keyboard=True,
                    input_field_placeholder="Отправьте URL фотографий",
                ).to_dict()
                send_telegram_message(
                    chat_id,
                    "Отправьте *URL* фотографий (через пробел или по одному):\n\n"
                    'Когда закончите, нажмите "✅ Завершить"',
                    reply_markup=rm,
                )
            elif text == "📷 Загрузить фото с устройства":
                state_data["photo_mode"] = "device"
                profile.telegram_state = state_data
                profile.save()
                rm = ReplyKeyboardMarkup(
                    [[KeyboardButton("✅ Завершить")], [KeyboardButton("❌ Отмена")]],
                    resize_keyboard=True,
                ).to_dict()
                send_telegram_message(
                    chat_id,
                    "Пришлите фотографии с устройства (до 6 штук):\n\n"
                    'Когда закончите, нажмите "✅ Завершить"',
                    reply_markup=rm,
                )
            elif text == "⏭️ Пропустить фото":
                send_telegram_message(
                    chat_id,
                    f"✅ Квартира успешно создана!\n\n"
                    f"Вы можете добавить фотографии позже.",
                )
                profile.telegram_state = {}
                profile.save()
                show_admin_menu(chat_id)
            else:
                send_telegram_message(
                    chat_id, "Пожалуйста, выберите способ загрузки фотографий."
                )
            return True

        # Завершение добавления фото
        if text == "✅ Завершить":
            photos_count = PropertyPhoto.objects.filter(property_id=prop_id).count()
            send_telegram_message(
                chat_id, f"✅ Квартира создана с {photos_count} фотографиями!"
            )
            profile.telegram_state = {}
            profile.save()
            show_admin_menu(chat_id)
            return True

        if text.startswith("➡️ Далее") or text.startswith("⬅️ Назад"):
            m = NAV_PAGE_RE.search(text)
            next_page = int(m.group(1)) if m else 1
            return show_admin_properties(chat_id, page=next_page)

        # Режим URL: обрабатываем текст со ссылками
        if photo_mode == 'url' and text and text not in ["✅ Завершить", "❌ Отмена"]:
            # Проверяем текущее количество фото перед добавлением
            current_count = PropertyPhoto.objects.filter(property_id=prop_id).count()

            if current_count >= 6:
                send_telegram_message(
                    chat_id,
                    "❌ *Достигнут максимум!*\n\n"
                    "Уже загружено 6 фотографий.\n"
                    "Нажмите «✅ Завершить» для сохранения."
                )
                return True

            urls = [u.strip() for u in text.split() if u.strip().startswith('http')]

            # Ограничиваем количество URL
            available_slots = 6 - current_count
            if len(urls) > available_slots:
                send_telegram_message(
                    chat_id,
                    f"⚠️ *Слишком много ссылок!*\n\n"
                    f"Можно добавить еще только {available_slots} фото.\n"
                    f"Отправьте не более {available_slots} ссылок."
                )
                return True

            created = 0
            for url in urls[:available_slots]:  # Ограничиваем количество
                try:
                    PropertyPhoto.objects.create(property_id=prop_id, image_url=url)
                    created += 1
                except Exception as e:
                    logger.warning(f"Bad URL {url}: {e}")

            if created > 0:
                total_photos = PropertyPhoto.objects.filter(property_id=prop_id).count()
                if total_photos >= 6:
                    send_telegram_message(
                        chat_id,
                        f"✅ *Максимум фотографий достигнут!*\n\n"
                        f"Загружено 6/6 фотографий.\n"
                        "Нажмите «✅ Завершить» для сохранения."
                    )
                else:
                    remaining = 6 - total_photos
                    send_telegram_message(
                        chat_id,
                        f"✅ Добавлено {created} фото.\n"
                        f"Всего: {total_photos}/6\n"
                        f"Можно добавить еще: {remaining}\n\n"
                        "Отправьте еще URL или нажмите «✅ Завершить»"
                    )
            else:
                send_telegram_message(
                    chat_id,
                    "❌ Не удалось добавить фотографии.\n"
                    "Проверьте корректность URL."
                )
            return True

        # Режим device: информируем что фото нужно отправлять не текстом
        if (
            photo_mode == "device"
            and text
            and text not in ["✅ Завершить", "❌ Отмена"]
        ):
            send_telegram_message(
                chat_id, "Пожалуйста, отправьте фотографии как изображения, а не текст."
            )
            return True

    return False


@log_handler
def quick_photo_management(chat_id, property_id):
    """Быстрый доступ к управлению фотографиями из списка квартир"""
    profile = _get_profile(chat_id)

    try:
        # Проверяем доступ к квартире
        if profile.role == 'admin':
            prop = Property.objects.get(id=property_id, owner=profile.user)
        else:  # super_admin
            prop = Property.objects.get(id=property_id)

        # Устанавливаем состояние редактирования
        profile.telegram_state = {
            'state': STATE_EDIT_PROPERTY_MENU,
            'editing_property_id': property_id
        }
        profile.save()

        # Запускаем управление фотографиями
        from .edit_handlers import handle_manage_photos_start
        handle_manage_photos_start(chat_id)

    except Property.DoesNotExist:
        send_telegram_message(chat_id, "❌ Квартира не найдена или у вас нет доступа.")

@log_handler
def handle_photo_upload(chat_id, update, context):
    """Обработка загружаемых фотографий с проверкой лимита в 6 штук."""
    profile = _get_profile(chat_id)
    state_data = profile.telegram_state or {}
    state = state_data.get('state')

    if state != STATE_ADMIN_ADD_PHOTOS:
        return False

    photo_mode = state_data.get('photo_mode')
    if photo_mode != 'device':
        return False

    prop_id = state_data['new_property'].get('id')
    if not prop_id:
        send_telegram_message(chat_id, "Ошибка: квартира не найдена.")
        return True

    # Проверяем текущее количество фото
    current_photos = PropertyPhoto.objects.filter(property_id=prop_id).count()

    # ИСПРАВЛЕНИЕ: Строгая проверка на 6 фото
    if current_photos >= 6:
        send_telegram_message(
            chat_id,
            "❌ *Достигнут максимум!*\n\n"
            "Можно загрузить максимум 6 фотографий.\n"
            "У вас уже загружено 6 фото.\n\n"
            "Нажмите «✅ Завершить» для сохранения."
        )
        return True

    # Проверяем, что не пытаются загрузить больше одной фотографии за раз
    if update.message and update.message.photo:
        photos = update.message.photo

        # Если пытаются загрузить несколько фото сразу (медиа-группа)
        # Telegram отправляет их по одной, но мы проверяем на всякий случай
        if len(photos) > 1 and (current_photos + 1) > 6:
            send_telegram_message(
                chat_id,
                f"⚠️ *Внимание!*\n\n"
                f"Вы можете загрузить еще {6 - current_photos} фото.\n"
                f"Отправляйте фотографии по одной."
            )
            return True

        created = 0
        bot = context.bot

        try:
            best_photo = max(photos, key=lambda p: getattr(p, 'file_size', 0) or 0)

            # Проверяем размер файла
            if hasattr(best_photo, 'file_size') and best_photo.file_size > 5 * 1024 * 1024:
                send_telegram_message(
                    chat_id,
                    "❌ *Фото слишком большое!*\n\n"
                    "Максимальный размер файла: 5 МБ.\n"
                    "Попробуйте уменьшить размер фото."
                )
                return True

            # Сохраняем фото
            file = bot.get_file(best_photo.file_id)

            import tempfile
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
            file.download(custom_path=tmp.name)

            with open(tmp.name, 'rb') as f:
                from django.core.files import File
                django_file = File(f, name=f"property_{prop_id}_{best_photo.file_id}.jpg")
                PropertyPhoto.objects.create(property_id=prop_id, image=django_file)

            import os
            os.unlink(tmp.name)
            created = 1

        except Exception as e:
            logger.error(f"Failed to save photo: {e}", exc_info=True)
            created = 0

        if created > 0:
            total_photos = PropertyPhoto.objects.filter(property_id=prop_id).count()

            # Проверяем достижение лимита после сохранения
            if total_photos >= 6:
                send_telegram_message(
                    chat_id,
                    f"✅ *Максимум фотографий загружен!*\n\n"
                    f"Загружено: 6/6 фотографий\n"
                    f"Нажмите «✅ Завершить» для сохранения квартиры."
                )
            else:
                remaining = 6 - total_photos
                send_telegram_message(
                    chat_id,
                    f"✅ *Фотография добавлена!*\n\n"
                    f"Загружено: {total_photos}/6\n"
                    f"Можно добавить еще: {remaining}\n\n"
                    f"Отправьте следующее фото или нажмите «✅ Завершить»"
                )
        else:
            send_telegram_message(
                chat_id,
                "❌ Не удалось сохранить фотографию.\n"
                "Попробуйте еще раз."
            )

        return True

    return False


@log_handler
def show_admin_menu(chat_id):
    """Показать главное админ-меню."""
    profile = _get_profile(chat_id)
    text = "🔧 *Административная панель*"
    keyboard = [
        [KeyboardButton("➕ Добавить квартиру")],
        [KeyboardButton("📊 Статистика")],
        [KeyboardButton("🏠 Мои квартиры")],
        [KeyboardButton("📈 Расширенная аналитика")],
    ]
    if profile.role in ("super_admin", "super_user"):
        keyboard.append([KeyboardButton("👥 Управление админами")])
        keyboard.append([KeyboardButton("📊 KO-фактор гостей")])
    keyboard.append([KeyboardButton("🧭 Главное меню")])
    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(
            keyboard, resize_keyboard=True, input_field_placeholder="Выберите действие"
        ).to_dict(),
    )


@log_handler
def show_admin_panel(chat_id):
    """Отобразить меню администратора."""
    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к админ‑панели.")
        return

    text = "🛠 *Панель администратора*.\nВыберите действие:"
    buttons = [
        [KeyboardButton("➕ Добавить квартиру"), KeyboardButton("🏠 Мои квартиры")],
        [KeyboardButton("📊 Статистика"), KeyboardButton("📝 Отзывы о гостях")],
        [
            KeyboardButton("📈 Экспорт XLSX"),
            KeyboardButton("🧭 Главное меню"),
        ]
    ]
    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(
            buttons, resize_keyboard=True, input_field_placeholder="Выберите действие"
        ).to_dict(),
    )

NAV_PAGE_RE = re.compile(r"\(стр\.?\s*(\d+)\)")


@log_handler
def show_admin_properties(chat_id, page: int = 1):
    """Показать список квартир админа с возможностью просмотра доступности (Reply + пагинация по 3)"""
    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    qs = (
        Property.objects.filter(owner=profile.user)
        if profile.role == "admin"
        else Property.objects.all()
    ).order_by("id")  # фиксируем порядок, чтобы страницы были стабильны

    total = qs.count()
    if total == 0:
        send_telegram_message(
            chat_id,
            "У вас пока нет квартир.",
            reply_markup=ReplyKeyboardMarkup(
                [
                    [KeyboardButton("🛠 Панель администратора")],
                    [KeyboardButton("🧭 Главное меню")],
                ],
                resize_keyboard=True,
            ).to_dict(),
        )
        return

    total_pages = max(1, ceil(total / PAGE_SIZE))
    page = max(1, min(page, total_pages))  # защита от выхода за границы

    start = (page - 1) * PAGE_SIZE
    page_props = qs[start: start + PAGE_SIZE]

    # Логируем для отладки
    logger.info(f"Admin properties: page={page}, total_pages={total_pages}, total={total}")

    lines = [
        "🏠 *Ваши квартиры:*\n",
        f"Страница {page}/{total_pages} • всего: {total}\n",
    ]
    keyboard = []

    for idx, prop in enumerate(page_props, start=start + 1):
        lines.append(
            f"{idx}. {prop.name}\n"
            f"   📍 {prop.district.city.name}, {prop.district.name}\n"
            f"   💰 {prop.price_per_day} ₸/сутки\n"
            f"   Статус: {prop.status}\n"
        )
        # остаёмся на Reply — две кнопки в строке
        keyboard.append([
            KeyboardButton(f"📊 Доступность #{prop.id}"),
            KeyboardButton(f"✏️ #{prop.id} {prop.name[:20]}"),  # Ограничиваем длину названия
        ])

    # Навигация (Reply-текст с номером целевой страницы)
    if total_pages > 1:
        nav_row = []
        if page > 1:
            nav_row.append(KeyboardButton(f"⬅️ Назад (стр. {page - 1})"))
        nav_row.append(KeyboardButton(f"📄 {page}/{total_pages}"))
        if page < total_pages:
            nav_row.append(KeyboardButton(f"➡️ Далее (стр. {page + 1})"))
        keyboard.append(nav_row)

        # Логируем кнопки навигации для отладки
        logger.info(f"Navigation buttons: {nav_row}")

    keyboard.append([KeyboardButton("🛠 Панель администратора")])
    keyboard.append([KeyboardButton("🧭 Главное меню")])

    text = "\n".join(lines)
    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


# Обработчик текста с Reply-кнопок
@log_handler
def handle_admin_properties_input(chat_id, text: str):
    """Обработчик ввода для страницы квартир админа"""
    # Пагинация
    if text.startswith("➡️ Далее") or text.startswith("⬅️ Назад"):
        import re
        match = re.search(r'стр\.\s*(\d+)', text)
        if match:
            next_page = int(match.group(1))
            logger.info(f"Navigating to page {next_page}")
            show_admin_properties(chat_id, page=next_page)
            return True
        else:
            logger.error(f"Could not parse page number from: {text}")
            send_telegram_message(chat_id, "❌ Ошибка навигации")
            return True

    # Информационная кнопка страницы
    if text.startswith("📄"):
        import re
        match = re.search(r'(\d+)/\d+', text)
        if match:
            current_page = int(match.group(1))
            show_admin_properties(chat_id, page=current_page)
            return True

    return False


@log_handler
def show_property_availability(chat_id, property_id):
    """Показать информацию о доступности квартиры (замена календаря)"""
    profile = _get_profile(chat_id)

    if profile.role not in ('admin', 'super_admin', 'super_user'):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    try:
        # Проверяем доступ к квартире
        if profile.role == 'admin':
            prop = Property.objects.get(id=property_id, owner=profile.user)
        else:
            prop = Property.objects.get(id=property_id)

        from datetime import date, timedelta
        from django.db.models import Q

        today = date.today()
        next_30_days = today + timedelta(days=30)

        # Получаем бронирования на ближайшие 30 дней
        bookings = Booking.objects.filter(
            property=prop,
            status__in=['confirmed', 'completed'],
            start_date__lt=next_30_days,
            end_date__gt=today
        ).order_by('start_date')

        text = (
            f"📊 *Информация о доступности*\n\n"
            f"🏠 {prop.name}\n"
            f"📅 Период: {today.strftime('%d.%m.%Y')} - {next_30_days.strftime('%d.%m.%Y')}\n"
            f"💰 Цена: {prop.price_per_day} ₸/сутки\n"
            f"📊 Статус: {prop.status}\n\n"
        )

        if not bookings.exists():
            text += "✅ *Квартира полностью свободна на ближайшие 30 дней*\n\n"
            text += f"💰 Потенциальный доход: {30 * prop.price_per_day:,.0f} ₸"
        else:
            text += "📋 *Занятые периоды:*\n"
            total_booked_days = 0
            total_revenue = 0

            for booking in bookings:
                guest_name = booking.user.get_full_name() or booking.user.username
                days = (min(booking.end_date, next_30_days) - max(booking.start_date, today)).days
                total_booked_days += days
                total_revenue += booking.total_price

                text += (
                    f"• {booking.start_date.strftime('%d.%m')} - "
                    f"{booking.end_date.strftime('%d.%m')} "
                    f"({days} дн.)\n"
                    f"  👤 {guest_name}\n"
                    f"  💰 {booking.total_price:,.0f} ₸\n\n"
                )

            # Статистика
            free_days = 30 - total_booked_days
            occupancy_rate = (total_booked_days / 30) * 100
            potential_revenue = 30 * prop.price_per_day

            text += f"📊 *Статистика на 30 дней:*\n"
            text += f"✅ Свободно: {free_days} дней\n"
            text += f"🏠 Занято: {total_booked_days} дней\n"
            text += f"📈 Загрузка: {occupancy_rate:.1f}%\n"
            text += f"💰 Доход: {total_revenue:,.0f} ₸\n"
            text += f"📊 Потенциал: {potential_revenue:,.0f} ₸\n"
            text += f"💸 Упущено: {potential_revenue - total_revenue:,.0f} ₸"

        # Кнопки
        keyboard = [
            [KeyboardButton("🏠 Мои квартиры")],
            [KeyboardButton("🧭 Главное меню")]
        ]

        send_telegram_message(
            chat_id,
            text,
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict()
        )

    except Property.DoesNotExist:
        send_telegram_message(chat_id, "Квартира не найдена.")


@log_handler
def handle_edit_property_choice(chat_id, text):
    """Обработка выбора в меню редактирования"""
    profile = _get_profile(chat_id)
    state_data = profile.telegram_state or {}
    property_id = state_data.get('editing_property_id')

    logger.info(f"handle_edit_property_choice: text='{text}', property_id={property_id}")

    if not property_id:
        send_telegram_message(chat_id, "Ошибка: квартира не найдена.")
        profile.telegram_state = {}
        profile.save()
        show_admin_properties(chat_id)
        return

    if text == "❌ Отмена":
        profile.telegram_state = {}
        profile.save()
        show_admin_properties(chat_id)
        return

    elif text == "💰 Изменить цену":
        state_data['state'] = STATE_WAITING_NEW_PRICE
        profile.telegram_state = state_data
        profile.save()

        keyboard = [[KeyboardButton("❌ Отмена")]]
        send_telegram_message(
            chat_id,
            "Введите новую цену за сутки (в тенге):",
            reply_markup=ReplyKeyboardMarkup(
                keyboard,
                resize_keyboard=True,
                input_field_placeholder="Например: 15000"
            ).to_dict()
        )

    elif text == "📝 Изменить описание":
        state_data['state'] = STATE_WAITING_NEW_DESCRIPTION
        profile.telegram_state = state_data
        profile.save()

        keyboard = [[KeyboardButton("❌ Отмена")]]
        send_telegram_message(
            chat_id,
            "Введите новое описание квартиры:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard,
                resize_keyboard=True,
                input_field_placeholder="Новое описание..."
            ).to_dict()
        )

    elif text == "📊 Изменить статус":
        state_data['state'] = STATE_WAITING_NEW_STATUS
        profile.telegram_state = state_data
        profile.save()

        keyboard = [
            [KeyboardButton("Свободна")],
            [KeyboardButton("На обслуживании")],
            [KeyboardButton("❌ Отмена")]
        ]
        send_telegram_message(
            chat_id,
            "Выберите новый статус:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict()
        )

    elif text == "📷 Управление фото":
        # ГЛАВНОЕ ИСПРАВЛЕНИЕ: правильный переход к управлению фото
        logger.info(f"Starting photo management for property {property_id}")

        # Обновляем состояние для управления фото
        state_data['state'] = STATE_PHOTO_MANAGEMENT
        profile.telegram_state = state_data
        profile.save()

        # Запускаем управление фотографиями
        from .edit_handlers import handle_manage_photos_start
        handle_manage_photos_start(chat_id)

    else:
        send_telegram_message(chat_id, "⚠️ Выберите действие из меню")
        # Показываем меню заново
        handle_edit_property_start(chat_id, property_id)


@log_handler
def show_detailed_statistics(chat_id, period="month"):
    """Показать детальную статистику и кнопки выбора периода."""
    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return
    start, today, label = _resolve_period_bounds(period)
    if profile.role == "admin":
        props = Property.objects.filter(owner=profile.user)
    else:
        props = Property.objects.all()
    bookings = Booking.objects.filter(
        property__in=props, created_at__gte=start, status__in=["confirmed", "completed"]
    )
    total_revenue = bookings.aggregate(Sum("total_price"))["total_price__sum"] or 0
    total_bookings = bookings.count()
    canceled = Booking.objects.filter(
        property__in=props, created_at__gte=start, status="cancelled"
    ).count()
    avg_value = total_revenue / total_bookings if total_bookings else 0
    # Текст
    text = (
        f"📊 *Статистика за {label}:*\n"
        f"Доход: {total_revenue:,.0f} ₸\n"
        f"Брони: {total_bookings}, Отменено: {canceled}\n"
        f"Средний чек: {avg_value:,.0f} ₸"
    )

    profile.telegram_state = {"state": "detailed_stats", "period": period}
    profile.save()

    buttons = _period_keyboard()
    buttons.append([KeyboardButton("📈 Экспорт XLSX")])
    buttons.append([KeyboardButton("🧭 Главное меню")])
    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(
            buttons, resize_keyboard=True, input_field_placeholder="Выберите действие"
        ).to_dict(),
    )


@log_handler
def show_extended_statistics(chat_id, period="month"):
    """Показать расширенную статистику для администратора."""
    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    start, today, period_label = _resolve_period_bounds(period)

    if profile.role == "admin":
        props = Property.objects.filter(owner=profile.user)
    else:
        props = Property.objects.all()

    if not props.exists():
        send_telegram_message(chat_id, "Для выбранного периода отсутствуют данные.")
        return

    base_filter = {
        "property__in": props,
        "created_at__gte": start,
    }

    bookings = Booking.objects.filter(
        status__in=["confirmed", "completed"], **base_filter
    )

    total_revenue = bookings.aggregate(total=Sum("total_price"))["total"] or 0
    total_bookings = bookings.count()
    canceled_qs = Booking.objects.filter(status="cancelled", **base_filter)
    canceled_count = canceled_qs.count()
    avg_check = total_revenue / total_bookings if total_bookings else 0

    duration_expr = ExpressionWrapper(
        F("end_date") - F("start_date"), output_field=DurationField()
    )
    lead_expr = ExpressionWrapper(
        F("start_date") - F("created_at"), output_field=DurationField()
    )
    annotated_bookings = bookings.annotate(
        duration_days=duration_expr, lead_days=lead_expr
    )

    total_nights_delta = annotated_bookings.aggregate(
        total=Sum("duration_days")
    )["total"]
    avg_stay_delta = annotated_bookings.aggregate(
        avg=Avg("duration_days")
    )["avg"]
    avg_lead_delta = annotated_bookings.aggregate(avg=Avg("lead_days"))["avg"]

    total_nights = total_nights_delta.days if total_nights_delta else 0
    avg_stay = avg_stay_delta.days if avg_stay_delta else 0
    avg_lead = avg_lead_delta.days if avg_lead_delta else 0

    period_days = max((today - start).days, 1)
    inventory = props.count() * period_days
    occupancy_rate = (total_nights / inventory * 100) if inventory else 0

    class_labels = {
        "comfort": "Комфорт",
        "business": "Бизнес",
        "premium": "Премиум",
    }
    class_revenue = bookings.values("property__property_class").annotate(
        total=Sum("total_price")
    )
    class_lines = [
        f"{class_labels.get(row['property__property_class'], row['property__property_class'])}: {row['total']:,.0f} ₸"
        for row in class_revenue
    ]

    top_props_revenue = (
        bookings.values("property__name")
        .annotate(total=Sum("total_price"))
        .order_by("-total")[:5]
    )
    top_props_count = (
        bookings.values("property__name")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )

    top_users_count = (
        bookings.values("user__username")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )
    top_users_spend = (
        bookings.values("user__username")
        .annotate(total=Sum("total_price"))
        .order_by("-total")[:5]
    )

    top_agents_revenue = (
        bookings.values("property__owner__username")
        .annotate(total=Sum("total_price"))
        .order_by("-total")[:5]
    )
    top_agents_count = (
        bookings.values("property__owner__username")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )

    reason_labels = dict(Booking.CANCEL_REASON_CHOICES)
    cancel_lines = [
        f"{reason_labels.get(row['cancel_reason'], row['cancel_reason'] or 'без причины')}: {row['total']}"
        for row in canceled_qs.values("cancel_reason").annotate(total=Count("id"))
        if row["total"]
    ]

    top_props_revenue_lines = [
        f"{idx}. {row['property__name']}: {row['total']:,.0f} ₸"
        for idx, row in enumerate(top_props_revenue, start=1)
    ]
    top_props_count_lines = [
        f"{idx}. {row['property__name']}: {row['count']}"
        for idx, row in enumerate(top_props_count, start=1)
    ]
    top_users_count_lines = [
        f"{idx}. {row['user__username']}: {row['count']}"
        for idx, row in enumerate(top_users_count, start=1)
    ]
    top_users_spend_lines = [
        f"{idx}. {row['user__username']}: {row['total']:,.0f} ₸"
        for idx, row in enumerate(top_users_spend, start=1)
    ]
    top_agents_revenue_lines = [
        f"{idx}. {row['property__owner__username']}: {row['total']:,.0f} ₸"
        for idx, row in enumerate(top_agents_revenue, start=1)
    ]
    top_agents_count_lines = [
        f"{idx}. {row['property__owner__username']}: {row['count']}"
        for idx, row in enumerate(top_agents_count, start=1)
    ]

    text_parts = [
        f"📈 *Расширенная аналитика за {period_label}:*",
        f"💰 Доход: {total_revenue:,.0f} ₸",
        f"📦 Брони: {total_bookings}, отмены: {canceled_count}",
        f"💳 Средний чек: {avg_check:,.0f} ₸",
        f"🏨 Занятость: {occupancy_rate:.1f}%",
        f"🛏️ Средняя длительность проживания: {avg_stay} ноч.",
        f"⏳ Средний срок бронирования до заезда: {avg_lead} дн.",
        "",
        "🏷️ Доход по классам:",
        *(class_lines or ["нет данных"]),
        "",
        "🏠 Топ-5 квартир по доходу:",
        *(top_props_revenue_lines or ["нет данных"]),
        "",
        "📊 Топ-5 квартир по количеству броней:",
        *(top_props_count_lines or ["нет данных"]),
        "",
        "👥 Топ-5 гостей по количеству заселений:",
        *(top_users_count_lines or ["нет данных"]),
        "",
        "💸 Топ-5 гостей по сумме трат:",
        *(top_users_spend_lines or ["нет данных"]),
        "",
        "🏢 Топ-5 риелторов по доходу:",
        *(top_agents_revenue_lines or ["нет данных"]),
        "",
        "📈 Топ-5 риелторов по количеству броней:",
        *(top_agents_count_lines or ["нет данных"]),
    ]

    if cancel_lines:
        text_parts.extend(["", "🚫 Отмены по причинам:", *cancel_lines])

    text = "\n".join(text_parts)

    profile.telegram_state = {"state": "extended_stats", "period": period}
    profile.save()

    buttons = _period_keyboard()
    buttons.append([KeyboardButton("📈 Экспорт XLSX")])
    buttons.append([KeyboardButton("🧭 Главное меню")])
    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(
            buttons, resize_keyboard=True, input_field_placeholder="Выберите период"
        ).to_dict(),
    )


@log_handler
def show_pending_guest_reviews(chat_id):
    """Показать список гостей, ожидающих отзыв"""
    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    # Находим завершенные бронирования без отзывов о госте
    from booking_bot.listings.models import GuestReview
    from datetime import date, timedelta

    # Брони за последние 30 дней
    cutoff_date = date.today() - timedelta(days=30)

    if profile.role == "admin":
        bookings = (
            Booking.objects.filter(
                property__owner=profile.user,
                status="completed",
                end_date__gte=cutoff_date,
            )
            .exclude(guest_review__isnull=False)
            .select_related("user", "property")[:10]
        )
    else:  # super_admin
        bookings = (
            Booking.objects.filter(status="completed", end_date__gte=cutoff_date)
            .exclude(guest_review__isnull=False)
            .select_related("user", "property")[:10]
        )

    if not bookings:
        text = "📝 Нет гостей, ожидающих отзыв."
        kb = [[KeyboardButton("🛠 Панель администратора")]]
    else:
        text = "📝 *Гости, ожидающие отзыв:*\n\n"
        kb = []

        for booking in bookings:
            guest_name = booking.user.get_full_name() or booking.user.username
            text += (
                f"• {guest_name}\n"
                f"  🏠 {booking.property.name}\n"
                f"  📅 {booking.start_date.strftime('%d.%m')} - {booking.end_date.strftime('%d.%m')}\n"
                f"  /review_guest_{booking.id}\n\n"
            )

        kb.append([KeyboardButton("🛠 Панель администратора")])

    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True).to_dict(),
    )


@log_handler
def handle_guest_review_start(chat_id, booking_id):
    """Начать процесс отзыва о госте"""
    profile = _get_profile(chat_id)

    try:
        if profile.role == "admin":
            booking = Booking.objects.get(
                id=booking_id, property__owner=profile.user, status="completed"
            )
        else:  # super_admin
            booking = Booking.objects.get(id=booking_id, status="completed")

        # Сохраняем в состояние
        profile.telegram_state = {
            "state": "guest_review_rating",
            "guest_review_booking_id": booking_id,
        }
        profile.save()

        guest_name = booking.user.get_full_name() or booking.user.username
        text = (
            f"📝 *Отзыв о госте*\n\n"
            f"Гость: {guest_name}\n"
            f"Квартира: {booking.property.name}\n"
            f"Период: {booking.start_date.strftime('%d.%m')} - {booking.end_date.strftime('%d.%m')}\n\n"
            "Оцените гостя от 1 до 5:"
        )

        kb = [
            [KeyboardButton("⭐"), KeyboardButton("⭐⭐"), KeyboardButton("⭐⭐⭐")],
            [KeyboardButton("⭐⭐⭐⭐"), KeyboardButton("⭐⭐⭐⭐⭐")],
            [KeyboardButton("❌ Отмена")],
        ]

        send_telegram_message(
            chat_id,
            text,
            reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True).to_dict(),
        )

    except Booking.DoesNotExist:
        send_telegram_message(chat_id, "Бронирование не найдено.")


@log_handler
def handle_guest_review_rating(chat_id, text):
    """Обработка рейтинга гостя"""
    profile = _get_profile(chat_id)

    # Подсчет звезд
    rating = text.count("⭐")
    if rating < 1 or rating > 5:
        send_telegram_message(chat_id, "Пожалуйста, выберите оценку от 1 до 5 звезд.")
        return

    sd = profile.telegram_state
    sd["guest_review_rating"] = rating
    sd["state"] = "guest_review_text"
    profile.telegram_state = sd
    profile.save()

    text = (
        f"Оценка: {'⭐' * rating}\n\n"
        "Напишите короткий комментарий о госте (или отправьте 'Пропустить'):"
    )

    kb = [[KeyboardButton("Пропустить")], [KeyboardButton("❌ Отмена")]]

    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(
            kb, resize_keyboard=True, input_field_placeholder="Ваш комментарий"
        ).to_dict(),
    )


@log_handler
def handle_guest_review_text(chat_id, text):
    """Сохранение отзыва о госте"""
    profile = _get_profile(chat_id)
    sd = profile.telegram_state

    booking_id = sd.get("guest_review_booking_id")
    rating = sd.get("guest_review_rating")

    if text == "Пропустить":
        text = ""

    try:
        booking = Booking.objects.get(id=booking_id)

        from booking_bot.listings.models import GuestReview

        GuestReview.objects.create(
            guest=booking.user,
            admin=profile.user,
            booking=booking,
            rating=rating,
            text=text,
        )

        # Обновляем KO-фактор гостя
        update_guest_ko_factor(booking.user)

        send_telegram_message(chat_id, "✅ Отзыв о госте сохранен!")

        # Очищаем состояние
        profile.telegram_state = {}
        profile.save()

        # Возврат в меню
        show_admin_panel(chat_id)

    except Exception as e:
        logger.error(f"Error saving guest review: {e}")
        send_telegram_message(chat_id, "❌ Ошибка при сохранении отзыва.")


def update_guest_ko_factor(user):
    """Обновить KO-фактор гостя на основе его истории"""
    from booking_bot.bookings.models import Booking
    from datetime import timedelta

    # Анализируем за последние 6 месяцев
    six_months_ago = date.today() - timedelta(days=180)

    total_bookings = Booking.objects.filter(
        user=user, created_at__gte=six_months_ago
    ).count()

    cancelled_bookings = Booking.objects.filter(
        user=user, created_at__gte=six_months_ago, status="cancelled", cancelled_by=user
    ).count()

    if total_bookings > 0:
        ko_factor = cancelled_bookings / total_bookings

        # Обновляем профиль
        profile = user.profile
        profile.ko_factor = ko_factor
        profile.save()

        logger.info(f"Updated KO-factor for {user.username}: {ko_factor:.2%}")


@log_handler
def show_top_users_statistics(chat_id):
    """Показать ТОП пользователей"""
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        send_telegram_message(chat_id, "❌ Нет доступа")
        return

    from django.db.models import Sum, Count
    from booking_bot.bookings.models import Booking

    # ТОП-5 по заселениям
    top_by_count = (
        Booking.objects.filter(status__in=["confirmed", "completed"])
        .values("user__username")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )

    # ТОП-5 по тратам
    top_by_sum = (
        Booking.objects.filter(status__in=["confirmed", "completed"])
        .values("user__username")
        .annotate(total=Sum("total_price"))
        .order_by("-total")[:5]
    )

    text = "👥 *ТОП пользователей*\n\n"
    text += "*По количеству заселений:*\n"
    for i, u in enumerate(top_by_count, 1):
        text += f"{i}. {u['user__username']}: {u['count']} броней\n"

    text += "\n*По сумме трат:*\n"
    for i, u in enumerate(top_by_sum, 1):
        text += f"{i}. {u['user__username']}: {u['total']:,.0f} ₸\n"

    send_telegram_message(chat_id, text)


@log_handler
def export_statistics_xlsx(chat_id: int, context=None, period: str = "month"):
    """Генерация и отправка XLSX отчета согласно текущему контексту аналитики."""

    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа.")
        return

    state_data = profile.telegram_state or {}
    scope = state_data.get("analytics_scope", "global")
    period = state_data.get("period", period)

    try:
        workbook = Workbook()
        worksheet = workbook.active

        if scope == "realtors":
            data, start, end, label = _collect_realtor_metrics(period)
            worksheet.title = "Realtors"
            worksheet.append(["#", "Риелтор", "Агентство", "Объектов", "Броней", "Доход, ₸"])
            for idx, item in enumerate(data, start=1):
                worksheet.append(
                    [
                        idx,
                        item["name"],
                        item["agency"] or "",
                        int(item["properties"] or 0),
                        int(item["bookings"] or 0),
                        float(item["revenue"] or 0),
                    ]
                )
        elif scope == "agency_list":
            data, start, end, label = _collect_agency_metrics(period)
            worksheet.title = "Agencies"
            worksheet.append(["#", "Агентство", "Риелторов", "Объектов", "Броней", "Доход, ₸"])
            for idx, item in enumerate(data, start=1):
                agency = item["agency"]
                worksheet.append(
                    [
                        idx,
                        agency.name,
                        int(item["members"] or 0),
                        int(item["properties"] or 0),
                        int(item["bookings"] or 0),
                        float(item["revenue"] or 0),
                    ]
                )
        elif scope == "agency_detail":
            agency_id = state_data.get("agency_id")
            if not agency_id:
                send_telegram_message(chat_id, "Выберите агентство перед экспортом.")
                return
            agency = RealEstateAgency.objects.filter(id=agency_id).first()
            if not agency:
                send_telegram_message(chat_id, "Агентство не найдено для экспорта.")
                return

            metrics = _collect_agency_detail_metrics(agency, period)
            summary = metrics.get("summary") or {"total_revenue": 0, "total_bookings": 0}

            worksheet.title = "Summary"
            worksheet.append(["Агентство", agency.name])
            worksheet.append(["Период", f"{metrics['start'].strftime('%d.%m.%Y')} – {metrics['end'].strftime('%d.%m.%Y')}"])
            worksheet.append(["Доход, ₸", float(summary.get("total_revenue") or 0)])
            worksheet.append(["Брони", int(summary.get("total_bookings") or 0)])
            worksheet.append(["Объектов", metrics["properties"]])
            worksheet.append(["Риелторов", metrics["members"]])
            worksheet.append(["Загрузка, %", metrics["occupancy"]])
            worksheet.append(["Отмен", metrics["cancelled_total"]])

            top_properties = workbook.create_sheet("Top Properties")
            top_properties.append(["#", "Квартира", "Доход, ₸", "Бронирований"])
            for idx, row in enumerate(metrics["top_props_revenue"], start=1):
                bookings_match = next((r for r in metrics["top_props_bookings"] if r["property__name"] == row["property__name"]), None)
                top_properties.append(
                    [
                        idx,
                        row["property__name"],
                        float(row.get("total") or 0),
                        int((bookings_match or {}).get("count") or 0),
                    ]
                )

            top_guests = workbook.create_sheet("Top Guests")
            top_guests.append(["#", "Гость", "Заселений", "Сумма, ₸"])
            for idx, row in enumerate(metrics["top_users_count"], start=1):
                spend_match = next((r for r in metrics["top_users_spend"] if r["user__username"] == row["user__username"]), None)
                top_guests.append(
                    [
                        idx,
                        row["user__username"],
                        int(row.get("count") or 0),
                        float((spend_match or {}).get("total") or 0),
                    ]
                )

            top_agents = workbook.create_sheet("Top Realtors")
            top_agents.append(["#", "Риелтор", "Бронирований", "Доход, ₸"])
            for idx, row in enumerate(metrics["top_realtor_bookings"], start=1):
                revenue_match = next(
                    (r for r in metrics["top_realtor_revenue"] if r["property__owner__username"] == row["property__owner__username"]),
                    None,
                )
                top_agents.append(
                    [
                        idx,
                        row["property__owner__username"],
                        int(row.get("count") or 0),
                        float((revenue_match or {}).get("total") or 0),
                    ]
                )

            cancellations = workbook.create_sheet("Cancellations")
            cancellations.append(["Причина", "Количество"])
            reason_labels = dict(Booking.CANCEL_REASON_CHOICES)
            for code, total in metrics["cancel_breakdown"]:
                cancellations.append([reason_labels.get(code, code or "Без причины"), total])
        else:
            start, end, label = _resolve_period_bounds(period)
            if profile.role == "admin":
                props = Property.objects.filter(owner=profile.user)
            else:
                props = Property.objects.all()

            bookings = Booking.objects.filter(
                property__in=props,
                created_at__gte=start,
                status__in=ANALYTIC_STATUSES,
            )

            worksheet.title = "Bookings"
            worksheet.append(["#", "Квартира", "Гость", "Заезд", "Выезд", "Сумма, ₸", "Статус"])
            for idx, booking in enumerate(bookings, start=1):
                worksheet.append(
                    [
                        idx,
                        booking.property.name,
                        booking.user.username if booking.user else "",
                        booking.start_date.strftime("%d.%m.%Y"),
                        booking.end_date.strftime("%d.%m.%Y"),
                        float(booking.total_price or 0),
                        booking.get_status_display(),
                    ]
                )

            summary_sheet = workbook.create_sheet("Summary")
            totals = bookings.aggregate(total=Sum("total_price"), count=Count("id"))
            summary_sheet.append(["Период", f"{start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}"])
            summary_sheet.append(["Доход, ₸", float(totals.get("total") or 0)])
            summary_sheet.append(["Брони", int(totals.get("count") or 0)])

        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    value = cell.value
                    cell.number_format = "#,##0" if isinstance(value, (int, float)) else cell.number_format
                    if value is not None:
                        max_length = max(max_length, len(str(value)))
                sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = f"analytics_{scope}_{period}.xlsx"
        caption = f"📈 Отчет за {period}"

        import requests

        bot_token = TELEGRAM_BOT_TOKEN
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        files = {
            "document": (
                filename,
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        data = {"chat_id": chat_id, "caption": caption}

        response = requests.post(url, data=data, files=files, timeout=30)
        if response.status_code != 200:
            send_telegram_message(chat_id, "Ошибка при отправке отчета")
    except Exception as exc:  # noqa: BLE001
        logger.exception("Failed to build analytics XLSX: %s", exc)
        send_telegram_message(chat_id, "Не удалось сформировать отчет. Попробуйте позже.")


@log_handler
def export_statistics_csv(chat_id: int, context=None, period: str = "month"):
    """Экспорт аналитики в CSV-файл согласно текущему выбранному скоупу."""

    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа.")
        return

    state_data = profile.telegram_state or {}
    scope = state_data.get("analytics_scope", "global")
    period = state_data.get("period", period)

    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)
    rows_written = 0

    def _write_section(title: str):
        nonlocal rows_written
        if rows_written:
            writer.writerow([])
            rows_written += 1
        writer.writerow([title])
        rows_written += 1

    def _write_table(headers, rows):
        nonlocal rows_written
        if headers:
            writer.writerow(headers)
            rows_written += 1
        for row in rows:
            writer.writerow(row)
            rows_written += 1

    try:
        if scope == "realtors":
            data, start, end, label = _collect_realtor_metrics(period)
            if not data:
                send_telegram_message(chat_id, "Нет данных для экспорта за выбранный период.")
                return

            _write_section(f"Риелторы — доход за {label}")
            _write_table(
                ["#", "Риелтор", "Агентство", "Объектов", "Броней", "Доход, ₸"],
                [
                    [
                        idx,
                        item["name"],
                        item["agency"] or "",
                        int(item["properties"] or 0),
                        int(item["bookings"] or 0),
                        float(item["revenue"] or 0),
                    ]
                    for idx, item in enumerate(data, start=1)
                ],
            )
            summary_period = start, end

        elif scope == "agency_list":
            data, start, end, label = _collect_agency_metrics(period)
            if not data:
                send_telegram_message(chat_id, "Нет данных по агентствам за выбранный период.")
                return

            _write_section(f"Агентства — результаты за {label}")
            _write_table(
                ["#", "Агентство", "Риелторов", "Объектов", "Броней", "Доход, ₸"],
                [
                    [
                        idx,
                        item["agency"].name,
                        int(item["members"] or 0),
                        int(item["properties"] or 0),
                        int(item["bookings"] or 0),
                        float(item["revenue"] or 0),
                    ]
                    for idx, item in enumerate(data, start=1)
                ],
            )
            summary_period = start, end

        elif scope == "agency_detail":
            agency_id = state_data.get("agency_id")
            if not agency_id:
                send_telegram_message(chat_id, "Сначала выберите агентство.")
                return

            agency = RealEstateAgency.objects.filter(id=agency_id).first()
            if not agency:
                send_telegram_message(chat_id, "Агентство не найдено.")
                return

            metrics = _collect_agency_detail_metrics(agency, period)
            summary = metrics.get("summary") or {}

            _write_section(f"Сводка по агентству {agency.name}")
            _write_table(
                ["Метрика", "Значение"],
                [
                    ("Период", f"{metrics['start'].strftime('%d.%m.%Y')} – {metrics['end'].strftime('%d.%m.%Y')}") ,
                    ("Доход, ₸", float(summary.get("total_revenue") or 0)),
                    ("Бронирований", int(summary.get("total_bookings") or 0)),
                    ("Объектов", metrics["properties"]),
                    ("Риелторов", metrics["members"]),
                    ("Загрузка, %", round(metrics["occupancy"], 2)),
                    ("Отмен", metrics["cancelled_total"]),
                ],
            )

            if metrics["cancel_breakdown"]:
                _write_section("Отмены по причинам")
                _write_table(["Причина", "Количество"], metrics["cancel_breakdown"])

            if metrics["top_props_revenue"]:
                _write_section("ТОП-5 квартир по доходу")
                _write_table(
                    ["#", "Квартира", "Доход, ₸"],
                    [
                        [idx, row["property__name"], float(row.get("total") or 0)]
                        for idx, row in enumerate(metrics["top_props_revenue"], start=1)
                    ],
                )

            if metrics["top_props_bookings"]:
                _write_section("ТОП-5 квартир по бронированиям")
                _write_table(
                    ["#", "Квартира", "Брони"],
                    [
                        [idx, row["property__name"], int(row.get("count") or 0)]
                        for idx, row in enumerate(metrics["top_props_bookings"], start=1)
                    ],
                )

            if metrics["top_users_count"]:
                _write_section("ТОП-5 гостей по количеству заселений")
                _write_table(
                    ["#", "Гость", "Заселений"],
                    [
                        [idx, row["user__username"], int(row.get("count") or 0)]
                        for idx, row in enumerate(metrics["top_users_count"], start=1)
                    ],
                )

            if metrics["top_users_spend"]:
                _write_section("ТОП-5 гостей по сумме трат")
                _write_table(
                    ["#", "Гость", "Сумма, ₸"],
                    [
                        [idx, row["user__username"], float(row.get("total") or 0)]
                        for idx, row in enumerate(metrics["top_users_spend"], start=1)
                    ],
                )

            if metrics["top_realtor_revenue"]:
                _write_section("ТОП-5 риелторов по доходу")
                _write_table(
                    ["#", "Риелтор", "Доход, ₸"],
                    [
                        [idx, row["property__owner__username"], float(row.get("total") or 0)]
                        for idx, row in enumerate(metrics["top_realtor_revenue"], start=1)
                    ],
                )

            if metrics["top_realtor_bookings"]:
                _write_section("ТОП-5 риелторов по бронированиям")
                _write_table(
                    ["#", "Риелтор", "Бронирований"],
                    [
                        [idx, row["property__owner__username"], int(row.get("count") or 0)]
                        for idx, row in enumerate(metrics["top_realtor_bookings"], start=1)
                    ],
                )

            summary_period = (metrics["start"], metrics["end"])

        else:
            start, end, label = _resolve_period_bounds(period)
            if profile.role == UserProfile.ROLE_ADMIN:
                properties_qs = Property.objects.filter(owner=profile.user)
            else:
                properties_qs = Property.objects.all()

            bookings = (
                Booking.objects.filter(
                    property__in=properties_qs,
                    created_at__gte=start,
                    created_at__lte=end,
                )
                .select_related("property", "user")
                .order_by("-created_at")
            )

            if not bookings.exists():
                send_telegram_message(chat_id, "За выбранный период бронирований не найдено.")
                return

            _write_section(f"Бронирования за {label}")
            _write_table(
                [
                    "ID",
                    "Квартира",
                    "Гость",
                    "Заезд",
                    "Выезд",
                    "Сумма, ₸",
                    "Статус",
                    "Создано",
                ],
                [
                    [
                        booking.id,
                        booking.property.name,
                        booking.user.get_full_name() or booking.user.username,
                        booking.start_date.strftime("%d.%m.%Y"),
                        booking.end_date.strftime("%d.%m.%Y"),
                        float(booking.total_price),
                        booking.get_status_display(),
                        booking.created_at.strftime("%d.%m.%Y %H:%M"),
                    ]
                    for booking in bookings
                ],
            )

            summary_period = (start, end)

        if rows_written == 0:
            send_telegram_message(chat_id, "Нет данных для экспорта.")
            return

        csv_content = csv_buffer.getvalue().encode("utf-8-sig")
        filename = f"analytics_{scope}_{period}.csv"
        caption_period = ""
        if summary_period:
            start_dt, end_dt = summary_period
            caption_period = f" {start_dt.strftime('%d.%m.%Y')} – {end_dt.strftime('%d.%m.%Y')}"
        caption = f"📈 CSV-отчет{caption_period}"

        import requests

        response = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument",
            data={"chat_id": chat_id, "caption": caption},
            files={
                "document": (
                    filename,
                    csv_content,
                    "text/csv",
                )
            },
            timeout=30,
        )

        if response.status_code != 200:
            send_telegram_message(chat_id, "Ошибка при отправке CSV-отчета")

    except Exception as exc:  # noqa: BLE001
        logger.exception("Failed to export analytics CSV: %s", exc)
        send_telegram_message(chat_id, "Не удалось сформировать CSV. Попробуйте позже.")


@log_handler
def show_property_management(chat_id, property_id):
    """Показать управление конкретной квартирой."""
    profile = _get_profile(chat_id)
    try:
        if profile.role == "admin":
            prop = Property.objects.get(id=property_id, owner=profile.user)
        else:
            prop = Property.objects.get(id=property_id)
    except Property.DoesNotExist:
        send_telegram_message(chat_id, "Квартира не найдена.")
        return
    # Собираем текст
    month = date.today() - timedelta(days=30)
    rev = (
        Booking.objects.filter(
            property=prop, created_at__gte=month, status__in=["confirmed", "completed"]
        ).aggregate(Sum("total_price"))["total_price__sum"]
        or 0
    )
    text = (
        f"🏠 *{prop.name}*\n"
        f"🛏 {prop.number_of_rooms} комн., {prop.area} м²\n"
        f"💰 {prop.price_per_day} ₸/сутки\n"
        f"Доход (30дн): {rev:,.0f} ₸"
    )
    buttons = [
        [KeyboardButton("Изменить цену")],
        [KeyboardButton("Изменить описание")],
        [KeyboardButton("Управление фото")],
        [KeyboardButton("🧭 Главное меню")],
    ]
    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(buttons, resize_keyboard=True).to_dict(),
    )


@log_handler
def show_super_admin_menu(chat_id):
    """Показать меню супер-админа"""
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    text = "👥 *Управление системой*\n\nВыберите действие:"

    buttons = [
        [KeyboardButton("➕ Добавить админа")],
        [KeyboardButton("📋 Список админов")],
        [KeyboardButton("❌ Удалить админа")],
        [KeyboardButton("📊 Статистика по городам")],
        [KeyboardButton("📊 Риелторы"), KeyboardButton("🏢 Агентства")],
        [KeyboardButton("📈 Общая статистика")],
        [KeyboardButton("🎯 План-факт")],
        [KeyboardButton("📊 KO-фактор гостей")],
        [KeyboardButton("🧭 Главное меню")],
    ]

    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(buttons, resize_keyboard=True).to_dict(),
    )


@log_handler
def handle_add_admin(chat_id):
    """Начать процесс добавления админа"""
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        return

    profile.telegram_state = {"state": "add_admin_username"}
    profile.save()

    keyboard = [[KeyboardButton("❌ Отмена")]]

    send_telegram_message(
        chat_id,
        "Введите username пользователя Telegram (без @) для назначения админом:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


@log_handler
def process_add_admin(chat_id, username):
    """Добавить админа по username"""
    try:
        # Ищем пользователя
        target_profile = (
            UserProfile.objects.filter(telegram_chat_id__isnull=False)
            .filter(user__username__iexact=f"telegram_{username}")
            .first()
        )

        if not target_profile:
            send_telegram_message(
                chat_id,
                f"❌ Пользователь с username {username} не найден.\n"
                "Он должен сначала запустить бота.",
            )
            return

        if target_profile.role == "admin":
            send_telegram_message(
                chat_id, "Этот пользователь уже является администратором"
            )
            return

        target_profile.role = "admin"
        target_profile.save()

        send_telegram_message(
            chat_id, f"✅ Пользователь {username} назначен администратором"
        )

        # Уведомляем нового админа
        if target_profile.telegram_chat_id:
            send_telegram_message(
                target_profile.telegram_chat_id,
                "🎉 Вы назначены администратором системы ЖильеGO!\n"
                "Теперь вам доступна панель администратора.",
            )

    except Exception as e:
        logger.error(f"Error adding admin: {e}")
        send_telegram_message(chat_id, "❌ Ошибка при добавлении админа")


@log_handler
def show_admins_list(chat_id):
    """Показать список администраторов"""
    admins = UserProfile.objects.filter(role="admin")

    if not admins.exists():
        send_telegram_message(chat_id, "Список администраторов пуст")
        return

    text = "👥 *Администраторы системы:*\n\n"

    for admin in admins:
        props_count = Property.objects.filter(owner=admin.user).count()
        username = admin.user.username.replace("telegram_", "@")
        text += f"• {username} - {props_count} объектов\n"

    keyboard = [[KeyboardButton("🧭 Главное меню")]]

    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


@log_handler
def show_city_statistics(chat_id, period="month"):
    """Показать статистику по городам для супер-админа"""
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        send_telegram_message(chat_id, "❌ Нет доступа")
        return

    from django.db.models import Sum, Count, Avg
    from datetime import date, timedelta

    # Определяем период
    today = date.today()
    if period == "week":
        start = today - timedelta(days=7)
    elif period == "month":
        start = today - timedelta(days=30)
    elif period == "quarter":
        start = today - timedelta(days=90)
    else:
        start = today - timedelta(days=365)

    # Собираем статистику по городам
    cities_data = []
    cities = City.objects.all()

    for city in cities:
        # Квартиры в городе
        city_properties = Property.objects.filter(district__city=city)

        # Бронирования за период
        city_bookings = Booking.objects.filter(
            property__district__city=city,
            created_at__gte=start,
            status__in=["confirmed", "completed"],
        )

        revenue = city_bookings.aggregate(Sum("total_price"))["total_price__sum"] or 0
        bookings_count = city_bookings.count()

        # Средняя загрузка
        total_nights = 0
        occupied_nights = 0

        for prop in city_properties:
            period_days = (today - start).days
            total_nights += period_days

            occupied = Booking.objects.filter(
                property=prop,
                status__in=["confirmed", "completed"],
                start_date__lte=today,
                end_date__gte=start,
            ).count()
            occupied_nights += occupied

        occupancy = (occupied_nights / total_nights * 100) if total_nights > 0 else 0

        # Средняя цена
        avg_price = (
            city_properties.aggregate(Avg("price_per_day"))["price_per_day__avg"] or 0
        )

        cities_data.append(
            {
                "name": city.name,
                "properties": city_properties.count(),
                "revenue": revenue,
                "bookings": bookings_count,
                "occupancy": occupancy,
                "avg_price": avg_price,
            }
        )

    # Сортируем по доходу
    cities_data.sort(key=lambda x: x["revenue"], reverse=True)

    # Формируем сообщение
    text = f"📊 *Статистика по городам за {period}*\n\n"

    for city in cities_data:
        text += (
            f"🏙 *{city['name']}*\n"
            f"• Объектов: {city['properties']}\n"
            f"• Доход: {city['revenue']:,.0f} ₸\n"
            f"• Бронирований: {city['bookings']}\n"
            f"• Загрузка: {city['occupancy']:.1f}%\n"
            f"• Средняя цена: {city['avg_price']:.0f} ₸\n\n"
        )

    # Общий итог
    total_revenue = sum(c["revenue"] for c in cities_data)
    total_bookings = sum(c["bookings"] for c in cities_data)

    text += (
        f"📈 *ИТОГО:*\n"
        f"Общий доход: {total_revenue:,.0f} ₸\n"
        f"Всего бронирований: {total_bookings}"
    )

    # Кнопки переключения периода
    keyboard = [
        [KeyboardButton("🏙 Неделя"), KeyboardButton("🏙 Месяц")],
        [KeyboardButton("🏙 Квартал"), KeyboardButton("🏙 Год")],
        [KeyboardButton("📥 Экспорт в CSV")],
        [KeyboardButton("🧭 Главное меню")],
    ]

    # Сохраняем состояние для переключения периодов
    profile.telegram_state = {"state": "city_stats", "period": period}
    profile.save()

    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


@log_handler
def show_plan_fact(chat_id):
    """Показать план-факт анализ"""
    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        send_telegram_message(chat_id, "❌ Нет доступа")
        return

    from booking_bot.listings.models import PropertyTarget
    from django.db.models import Sum
    from datetime import date
    import calendar

    # Текущий месяц
    today = date.today()
    month_start = date(today.year, today.month, 1)

    # Определяем квартиры для анализа
    if profile.role == "admin":
        properties = Property.objects.filter(owner=profile.user)
    else:
        properties = Property.objects.all()

    text = f"🎯 *План-факт за {calendar.month_name[today.month]} {today.year}*\n\n"

    total_plan_revenue = 0
    total_fact_revenue = 0

    for prop in properties[:10]:  # Ограничиваем 10 объектами
        # Получаем цель
        try:
            target = PropertyTarget.objects.get(property=prop, month=month_start)
            plan_revenue = target.target_revenue
            plan_occupancy = target.target_occupancy
        except PropertyTarget.DoesNotExist:
            # Если цели нет, ставим по умолчанию
            days_in_month = calendar.monthrange(today.year, today.month)[1]
            plan_revenue = prop.price_per_day * days_in_month * 0.6  # 60% загрузка
            plan_occupancy = 60

        # Факт
        fact_bookings = Booking.objects.filter(
            property=prop,
            created_at__month=today.month,
            created_at__year=today.year,
            status__in=["confirmed", "completed"],
        )

        fact_revenue = (
            fact_bookings.aggregate(Sum("total_price"))["total_price__sum"] or 0
        )

        # Расчет загрузки
        days_passed = today.day
        occupied_days = 0

        for booking in fact_bookings:
            if booking.start_date.month == today.month:
                days = min((booking.end_date - booking.start_date).days, days_passed)
                occupied_days += days

        fact_occupancy = (occupied_days / days_passed * 100) if days_passed > 0 else 0

        # Выполнение плана
        revenue_completion = (
            (fact_revenue / plan_revenue * 100) if plan_revenue > 0 else 0
        )

        # Эмодзи статуса
        if revenue_completion >= 100:
            status_emoji = "✅"
        elif revenue_completion >= 70:
            status_emoji = "⚠️"
        else:
            status_emoji = "❌"

        text += (
            f"{status_emoji} *{prop.name}*\n"
            f"План: {plan_revenue:,.0f} ₸ | Факт: {fact_revenue:,.0f} ₸\n"
            f"Выполнение: {revenue_completion:.0f}%\n"
            f"Загрузка: {fact_occupancy:.0f}% (план {plan_occupancy:.0f}%)\n\n"
        )

        total_plan_revenue += plan_revenue
        total_fact_revenue += fact_revenue

    # Итоги
    total_completion = (
        (total_fact_revenue / total_plan_revenue * 100) if total_plan_revenue > 0 else 0
    )

    text += (
        f"📊 *ИТОГО:*\n"
        f"План: {total_plan_revenue:,.0f} ₸\n"
        f"Факт: {total_fact_revenue:,.0f} ₸\n"
        f"Выполнение: {total_completion:.0f}%"
    )

    keyboard = [
        [KeyboardButton("🎯 Установить цели")],
        [KeyboardButton("🧭 Главное меню")],
    ]

    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


@log_handler
def set_property_targets(chat_id):
    """Начать процесс установки целей"""
    profile = _get_profile(chat_id)

    if profile.role == "admin":
        properties = Property.objects.filter(owner=profile.user)
    else:
        properties = Property.objects.all()

    if not properties.exists():
        send_telegram_message(chat_id, "У вас нет объектов для установки целей")
        return

    # Показываем список объектов
    keyboard = []
    for prop in properties[:10]:
        keyboard.append([KeyboardButton(f"Цель для {prop.id}: {prop.name[:30]}")])

    keyboard.append([KeyboardButton("❌ Отмена")])

    profile.telegram_state = {"state": "select_property_for_target"}
    profile.save()

    send_telegram_message(
        chat_id,
        "Выберите объект для установки целей:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


@log_handler
def handle_target_property_selection(chat_id, text):
    """Обработка выбора объекта для целей"""
    import re

    match = re.search(r"Цель для (\d+):", text)

    if not match:
        send_telegram_message(chat_id, "Неверный выбор")
        return

    property_id = int(match.group(1))
    profile = _get_profile(chat_id)

    profile.telegram_state = {
        "state": "set_target_revenue",
        "target_property_id": property_id,
    }
    profile.save()

    keyboard = [[KeyboardButton("❌ Отмена")]]

    send_telegram_message(
        chat_id,
        "Введите целевую выручку на месяц (в тенге):",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


@log_handler
def save_property_target(chat_id, revenue_text):
    """Сохранить цель для объекта"""
    try:
        revenue = float(revenue_text.replace(",", "").replace(" ", ""))
    except ValueError:
        send_telegram_message(chat_id, "Неверный формат суммы")
        return

    profile = _get_profile(chat_id)
    sd = profile.telegram_state or {}
    property_id = sd.get("target_property_id")

    if not property_id:
        return

    from booking_bot.listings.models import PropertyTarget
    from datetime import date

    month_start = date(date.today().year, date.today().month, 1)

    PropertyTarget.objects.update_or_create(
        property_id=property_id,
        month=month_start,
        defaults={
            "target_revenue": revenue,
            "target_occupancy": 60,  # По умолчанию 60%
        },
    )

    send_telegram_message(chat_id, f"✅ Цель установлена: {revenue:,.0f} ₸/месяц")

    profile.telegram_state = {}
    profile.save()
    show_plan_fact(chat_id)


@log_handler
def handle_remove_admin(chat_id):
    """Начать процесс удаления админа"""
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        return

    admins = UserProfile.objects.filter(role="admin")

    if not admins.exists():
        send_telegram_message(chat_id, "Нет администраторов для удаления")
        return

    keyboard = []
    for admin in admins:
        username = admin.user.username.replace("telegram_", "")
        keyboard.append([KeyboardButton(f"Удалить {username}")])

    keyboard.append([KeyboardButton("❌ Отмена")])

    profile.telegram_state = {"state": "remove_admin"}
    profile.save()

    send_telegram_message(
        chat_id,
        "Выберите администратора для удаления:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


@log_handler
def process_remove_admin(chat_id, text):
    """Удалить админа"""
    if text.startswith("Удалить "):
        username = text.replace("Удалить ", "")

        try:
            target_profile = UserProfile.objects.get(
                user__username=f"telegram_{username}", role="admin"
            )

            target_profile.role = "user"
            target_profile.save()

            send_telegram_message(
                chat_id, f"✅ Пользователь {username} больше не администратор"
            )

            # Уведомляем бывшего админа
            if target_profile.telegram_chat_id:
                send_telegram_message(
                    target_profile.telegram_chat_id,
                    "Ваши административные права отозваны.",
                )

        except UserProfile.DoesNotExist:
            send_telegram_message(chat_id, "Администратор не найден")


@log_handler
def prompt_guest_review(chat_id, booking_id):
    """Запрос отзыва об госте от админа"""
    profile = _get_profile(chat_id)
    if profile.role not in ("admin", "super_admin", "super_user"):
        return

    try:
        booking = Booking.objects.get(id=booking_id)

        # Проверяем, что это квартира админа
        if booking.property.owner != profile.user and profile.role not in ("super_admin", "super_user"):
            return

        # Проверяем, что нет отзыва
        from booking_bot.listings.models import GuestReview

        if GuestReview.objects.filter(booking=booking).exists():
            return

        profile.telegram_state = {
            "state": "admin_guest_review",
            "review_booking_id": booking_id,
        }
        profile.save()

        text = (
            f"📝 *Оставьте отзыв о госте*\n\n"
            f"Гость: {booking.user.first_name} {booking.user.last_name}\n"
            f"Квартира: {booking.property.name}\n"
            f"Период: {booking.start_date.strftime('%d.%m')} - {booking.end_date.strftime('%d.%m.%Y')}\n\n"
            "Оцените гостя от 1 до 5:"
        )

        keyboard = [
            [KeyboardButton("1⭐"), KeyboardButton("2⭐"), KeyboardButton("3⭐")],
            [KeyboardButton("4⭐"), KeyboardButton("5⭐")],
            [KeyboardButton("❌ Пропустить")],
        ]

        send_telegram_message(
            chat_id,
            text,
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
        )

    except Booking.DoesNotExist:
        pass


@log_handler
def handle_guest_review_rating(chat_id, text):
    """Обработка рейтинга гостя"""
    profile = _get_profile(chat_id)
    sd = profile.telegram_state or {}

    if text == "❌ Пропустить":
        profile.telegram_state = {}
        profile.save()
        show_admin_menu(chat_id)
        return

    # Извлекаем рейтинг
    if "⭐" in text:
        rating = int(text[0])
        sd["guest_rating"] = rating
        sd["state"] = "admin_guest_review_text"
        profile.telegram_state = sd
        profile.save()

        keyboard = [[KeyboardButton("Без комментария")], [KeyboardButton("❌ Отмена")]]

        send_telegram_message(
            chat_id,
            f"Оценка: {rating}⭐\n\nДобавьте комментарий или нажмите 'Без комментария':",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
        )


@log_handler
def save_guest_review(chat_id, text):
    """Сохранение отзыва о госте"""
    profile = _get_profile(chat_id)
    sd = profile.telegram_state or {}

    booking_id = sd.get("review_booking_id")
    rating = sd.get("guest_rating")

    if text == "Без комментария":
        text = ""

    try:
        booking = Booking.objects.get(id=booking_id)
        from booking_bot.listings.models import GuestReview

        GuestReview.objects.create(
            booking=booking,
            reviewer=profile.user,
            guest=booking.user,
            rating=rating,
            text=text,
        )

        send_telegram_message(chat_id, "✅ Отзыв о госте сохранен")

        # Обновляем KO-фактор если нужно
        guest_profile = booking.user.profile
        # Логика подсчета рейтинга гостя
        avg_rating = GuestReview.objects.filter(guest=booking.user).aggregate(
            Avg("rating")
        )["rating__avg"]

        if avg_rating and avg_rating < 3:  # Низкий рейтинг
            guest_profile.ko_factor = 0.7  # Повышаем KO-фактор
            guest_profile.save()

    except Exception as e:
        logger.error(f"Error saving guest review: {e}")

    profile.telegram_state = {}
    profile.save()
    show_admin_menu(chat_id)


@log_handler
def show_ko_factor_report(chat_id):
    """Показать отчет по KO-фактору гостей"""
    profile = _get_profile(chat_id)
    if profile.role not in ("super_admin", "super_user"):
        send_telegram_message(chat_id, "❌ Нет доступа")
        return

    from django.db.models import Count, Q

    # Получаем пользователей с высоким KO-фактором
    users_with_bookings = (
        User.objects.filter(bookings__isnull=False)
        .annotate(
            total_bookings=Count("bookings"),
            cancelled_bookings=Count(
                "bookings",
                filter=Q(bookings__status="cancelled", bookings__cancelled_by=F("id")),
            ),
        )
        .filter(total_bookings__gte=3)  # Минимум 3 бронирования
    )

    high_ko_users = []

    for user in users_with_bookings:
        if user.cancelled_bookings > 0:
            ko_factor = (user.cancelled_bookings / user.total_bookings) * 100
            if ko_factor > 30:  # Показываем с KO > 30%
                high_ko_users.append(
                    {
                        "user": user,
                        "ko_factor": ko_factor,
                        "total": user.total_bookings,
                        "cancelled": user.cancelled_bookings,
                    }
                )

    # Сортируем по KO-фактору
    high_ko_users.sort(key=lambda x: x["ko_factor"], reverse=True)

    text = "📊 *KO-фактор гостей*\n\n"

    if not high_ko_users:
        text += "Нет гостей с высоким процентом отмен"
    else:
        for data in high_ko_users[:15]:  # Топ-15
            user = data["user"]
            emoji = "🔴" if data["ko_factor"] > 50 else "🟡"

            text += (
                f"{emoji} {user.first_name} {user.last_name}\n"
                f"KO: {data['ko_factor']:.0f}% "
                f"({data['cancelled']}/{data['total']} отмен)\n"
            )

            if data["ko_factor"] > 50:
                text += "⚠️ Требуется предоплата\n"

            text += "\n"

    keyboard = [
        [KeyboardButton("📥 Экспорт KO-факторов")],
        [KeyboardButton("🧭 Главное меню")],
    ]

    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict(),
    )


# Добавить в файл booking_bot/telegram_bot/admin_handlers.py

# Новые состояния для редактирования
STATE_EDIT_PROPERTY_MENU = 'edit_property_menu'
STATE_EDIT_PROPERTY_PRICE = 'edit_property_price'
STATE_EDIT_PROPERTY_DESC = 'edit_property_desc'
STATE_EDIT_PROPERTY_STATUS = 'edit_property_status'


@log_handler
def handle_edit_property_start(chat_id, property_id):
    """Начать процесс редактирования квартиры"""
    profile = _get_profile(chat_id)

    try:
        # Проверяем доступ к квартире
        if profile.role == 'admin':
            prop = Property.objects.get(id=property_id, owner=profile.user)
        else:  # super_admin
            prop = Property.objects.get(id=property_id)

        # Сохраняем в состояние
        profile.telegram_state = {
            'state': STATE_EDIT_PROPERTY_MENU,
            'editing_property_id': property_id
        }
        profile.save()

        # Показываем текущую информацию и меню редактирования
        text = (
            f"✏️ *Редактирование квартиры*\n\n"
            f"🏠 {prop.name}\n"
            f"📝 {prop.description[:100]}...\n"
            f"💰 Текущая цена: {prop.price_per_day} ₸/сутки\n"
            f"📊 Статус: {prop.status}\n\n"
            "Что хотите изменить?"
        )

        keyboard = [
            [KeyboardButton("💰 Изменить цену")],
            [KeyboardButton("📝 Изменить описание")],
            [KeyboardButton("📊 Изменить статус")],
            [KeyboardButton("📷 Управление фото")],
            [KeyboardButton("❌ Отмена")]
        ]

        send_telegram_message(
            chat_id,
            text,
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict()
        )

    except Property.DoesNotExist:
        send_telegram_message(chat_id, "Квартира не найдена или у вас нет доступа.")


@log_handler
def handle_edit_property_menu(chat_id, text):
    """Обработка выбора в меню редактирования"""
    profile = _get_profile(chat_id)
    state_data = profile.telegram_state or {}
    property_id = state_data.get('editing_property_id')

    if not property_id:
        send_telegram_message(chat_id, "Ошибка: квартира не найдена.")
        return

    if text == "❌ Отмена":
        profile.telegram_state = {}
        profile.save()
        show_admin_properties(chat_id)
        return

    elif text == "💰 Изменить цену":
        state_data['state'] = STATE_EDIT_PROPERTY_PRICE
        profile.telegram_state = state_data
        profile.save()

        keyboard = [[KeyboardButton("❌ Отмена")]]
        send_telegram_message(
            chat_id,
            "Введите новую цену за сутки (в тенге):",
            reply_markup=ReplyKeyboardMarkup(
                keyboard,
                resize_keyboard=True,
                input_field_placeholder="Например: 15000"
            ).to_dict()
        )

    elif text == "📝 Изменить описание":
        state_data['state'] = STATE_EDIT_PROPERTY_DESC
        profile.telegram_state = state_data
        profile.save()

        keyboard = [[KeyboardButton("❌ Отмена")]]
        send_telegram_message(
            chat_id,
            "Введите новое описание квартиры:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard,
                resize_keyboard=True,
                input_field_placeholder="Новое описание..."
            ).to_dict()
        )

    elif text == "📊 Изменить статус":
        state_data['state'] = STATE_EDIT_PROPERTY_STATUS
        profile.telegram_state = state_data
        profile.save()

        keyboard = [
            [KeyboardButton("Свободна")],
            [KeyboardButton("На обслуживании")],
            [KeyboardButton("❌ Отмена")]
        ]
        send_telegram_message(
            chat_id,
            "Выберите новый статус:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True).to_dict()
        )

    elif text == "📷 Управление фото":
        send_telegram_message(
            chat_id,
            "Управление фотографиями пока в разработке.\n"
            "Используйте веб-панель для изменения фотографий."
        )


@log_handler
def handle_edit_property_price(chat_id, text):
    """Обработка изменения цены"""
    profile = _get_profile(chat_id)
    state_data = profile.telegram_state or {}
    property_id = state_data.get('editing_property_id')

    if text == "❌ Отмена":
        handle_edit_property_start(chat_id, property_id)
        return

    try:
        new_price = float(text.replace(',', '.'))
        if new_price <= 0:
            raise ValueError("Цена должна быть положительной")

        # Обновляем цену
        prop = Property.objects.get(id=property_id)
        old_price = prop.price_per_day
        prop.price_per_day = new_price
        prop.save()

        send_telegram_message(
            chat_id,
            f"✅ Цена успешно изменена!\n"
            f"Было: {old_price} ₸\n"
            f"Стало: {new_price} ₸"
        )

        # Возвращаемся в меню редактирования
        handle_edit_property_start(chat_id, property_id)

    except ValueError:
        send_telegram_message(chat_id, "Неверный формат цены. Введите число.")
    except Property.DoesNotExist:
        send_telegram_message(chat_id, "Квартира не найдена.")


@log_handler
def handle_edit_property_desc(chat_id, text):
    """Обработка изменения описания"""
    profile = _get_profile(chat_id)
    state_data = profile.telegram_state or {}
    property_id = state_data.get('editing_property_id')

    if text == "❌ Отмена":
        handle_edit_property_start(chat_id, property_id)
        return

    try:
        # Обновляем описание
        prop = Property.objects.get(id=property_id)
        prop.description = text.strip()
        prop.save()

        send_telegram_message(
            chat_id,
            "✅ Описание успешно обновлено!"
        )

        # Возвращаемся в меню редактирования
        handle_edit_property_start(chat_id, property_id)

    except Property.DoesNotExist:
        send_telegram_message(chat_id, "Квартира не найдена.")


@log_handler
def handle_edit_property_status(chat_id, text):
    """Обработка изменения статуса"""
    profile = _get_profile(chat_id)
    state_data = profile.telegram_state or {}
    property_id = state_data.get('editing_property_id')

    if text == "❌ Отмена":
        handle_edit_property_start(chat_id, property_id)
        return

    if text not in ["Свободна", "На обслуживании"]:
        send_telegram_message(chat_id, "Выберите статус из предложенных вариантов.")
        return

    try:
        # Обновляем статус
        prop = Property.objects.get(id=property_id)
        old_status = prop.status
        prop.status = text
        prop.save()

        send_telegram_message(
            chat_id,
            f"✅ Статус успешно изменен!\n"
            f"Было: {old_status}\n"
            f"Стало: {text}"
        )

        # Возвращаемся в меню редактирования
        handle_edit_property_start(chat_id, property_id)

    except Property.DoesNotExist:
        send_telegram_message(chat_id, "Квартира не найдена.")


# Добавить в файл booking_bot/telegram_bot/admin_handlers.py

# Состояния для модерации отзывов
STATE_MODERATE_REVIEWS = 'moderate_reviews'
STATE_MODERATE_REVIEW_ACTION = 'moderate_review_action'


@log_handler
def show_pending_reviews(chat_id):
    """Показать список неодобренных отзывов для модерации"""
    profile = _get_profile(chat_id)
    if profile.role not in ('admin', 'super_admin', 'super_user'):
        send_telegram_message(chat_id, "У вас нет доступа к этой функции.")
        return

    from booking_bot.listings.models import Review

    # Получаем неодобренные отзывы
    if profile.role == 'admin':
        # Админ видит только отзывы о своих квартирах
        pending_reviews = Review.objects.filter(
            property__owner=profile.user,
            is_approved=False
        ).select_related('property', 'user').order_by('-created_at')[:10]
    else:
        # Супер-админ видит все
        pending_reviews = Review.objects.filter(
            is_approved=False
        ).select_related('property', 'user').order_by('-created_at')[:10]

    if not pending_reviews:
        text = "📝 Нет отзывов, ожидающих модерации."
        kb = [[KeyboardButton("🛠 Панель администратора")]]
    else:
        text = "📝 *Отзывы на модерации:*\n\n"
        kb = []

        for review in pending_reviews:
            guest_name = review.user.get_full_name() or review.user.username
            text += (
                f"• ID: {review.id}\n"
                f"  Гость: {guest_name}\n"
                f"  Квартира: {review.property.name}\n"
                f"  Оценка: {'⭐' * review.rating}\n"
                f"  Текст: {review.comment[:100]}...\n"
                f"  /moderate_{review.id}\n\n"
            )

        kb.append([KeyboardButton("🛠 Панель администратора")])

    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True).to_dict()
    )


@log_handler
def handle_moderate_review_start(chat_id, review_id):
    """Начать модерацию конкретного отзыва"""
    profile = _get_profile(chat_id)

    try:
        from booking_bot.listings.models import Review

        if profile.role == 'admin':
            review = Review.objects.get(
                id=review_id,
                property__owner=profile.user,
                is_approved=False
            )
        else:
            review = Review.objects.get(
                id=review_id,
                is_approved=False
            )

        # Сохраняем в состояние
        profile.telegram_state = {
            'state': STATE_MODERATE_REVIEW_ACTION,
            'moderating_review_id': review_id
        }
        profile.save()

        guest_name = review.user.get_full_name() or review.user.username
        text = (
            f"📝 *Модерация отзыва #{review_id}*\n\n"
            f"Гость: {guest_name}\n"
            f"Квартира: {review.property.name}\n"
            f"Оценка: {'⭐' * review.rating}\n"
            f"Дата: {review.created_at.strftime('%d.%m.%Y')}\n\n"
            f"*Текст отзыва:*\n{review.comment}\n\n"
            "Что сделать с отзывом?"
        )

        kb = [
            [KeyboardButton("✅ Одобрить")],
            [KeyboardButton("❌ Отклонить")],
            [KeyboardButton("🔙 Назад к списку")]
        ]

        send_telegram_message(
            chat_id,
            text,
            reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True).to_dict()
        )

    except Review.DoesNotExist:
        send_telegram_message(chat_id, "Отзыв не найден или уже обработан.")


@log_handler
def handle_moderate_review_action(chat_id, text):
    """Обработка действия модерации"""
    profile = _get_profile(chat_id)
    state_data = profile.telegram_state or {}
    review_id = state_data.get('moderating_review_id')

    if not review_id:
        send_telegram_message(chat_id, "Ошибка: отзыв не найден.")
        return

    from booking_bot.listings.models import Review

    try:
        review = Review.objects.get(id=review_id)

        if text == "✅ Одобрить":
            review.is_approved = True
            review.save()

            send_telegram_message(
                chat_id,
                f"✅ Отзыв #{review_id} одобрен и теперь виден пользователям."
            )

            # Уведомляем автора отзыва
            if hasattr(review.user, 'profile') and review.user.profile.telegram_chat_id:
                send_telegram_message(
                    review.user.profile.telegram_chat_id,
                    f"✅ Ваш отзыв о квартире {review.property.name} был одобрен!"
                )

        elif text == "❌ Отклонить":
            # Удаляем отклоненный отзыв
            review.delete()

            send_telegram_message(
                chat_id,
                f"❌ Отзыв #{review_id} отклонен и удален."
            )

        elif text == "🔙 Назад к списку":
            show_pending_reviews(chat_id)
            return
        else:
            send_telegram_message(chat_id, "Выберите действие из предложенных.")
            return

        # Очищаем состояние и возвращаемся к списку
        profile.telegram_state = {}
        profile.save()
        show_pending_reviews(chat_id)

    except Review.DoesNotExist:
        send_telegram_message(chat_id, "Отзыв не найден.")
        profile.telegram_state = {}
        profile.save()


# Добавить в show_admin_panel функцию кнопку модерации
@log_handler
def show_admin_panel_with_moderation(chat_id):
    """Отобразить меню администратора с модерацией отзывов."""
    profile = _get_profile(chat_id)
    if profile.role not in ('admin', 'super_admin', 'super_user'):
        send_telegram_message(chat_id, "У вас нет доступа к админ‑панели.")
        return

    text = "🛠 *Панель администратора*.\nВыберите действие:"
    buttons = [
        [KeyboardButton("➕ Добавить квартиру"), KeyboardButton("🏠 Мои квартиры")],
        [KeyboardButton("📊 Статистика"), KeyboardButton("📈 Экспорт XLSX")],
        [KeyboardButton("📝 Отзывы о гостях"), KeyboardButton("✅ Модерация отзывов")],
        [KeyboardButton("🧭 Главное меню")]
    ]
    send_telegram_message(
        chat_id,
        text,
        reply_markup=ReplyKeyboardMarkup(
            buttons, resize_keyboard=True,
            input_field_placeholder="Выберите действие"
        ).to_dict()
    )
